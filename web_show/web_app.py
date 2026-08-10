"""
人脸识别考勤系统 - Web 交互界面（苹果风格 · 白蓝主题 · 分步向导）

流程：
  第1步 人脸识别：上传 Ref_Figure / Target_Figure 图片 → 点按钮识别 → 生成 output/00_程序人脸识别结果.xlsx
  第2步 上传输入表：5 个 input/ 输入表，全部就绪后进入下一步
  第3步 生成表格：表1~表10 依次生成，生成 05/06 时弹窗人工确认（可回传修正表）
  第4步 完成下载：下载 08/09/10 单表 + 一键 ZIP

启动（必须用 AI_PY312 环境，该环境有 insightface/cv2）：
  C:/Users/29037/.conda/envs/AI_PY312/python.exe web_show/web_app.py
"""

import io
import os
import sys
import time
import contextlib
import threading
import traceback
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 注入项目根目录，便于 import face_app
BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

WEB_SHOW_DIR = Path(__file__).resolve().parent

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

from flask import Flask, request, jsonify, send_file, send_from_directory

from face_app import (  # 仅依赖 pandas/openpyxl，无需 insightface（懒加载）
    get_table1, get_table2, get_table3, get_table4,
    get_table5, get_table6, get_table7, get_table8_9_10,
)
from face_app.generate_table import (  # 表格输出路径的单一数据源
    TABLE1_FILE, TABLE2_FILE, TABLE3_FILE, TABLE4_FILE,
    TABLE5_FILE, TABLE5_ERROR_FILE, TABLE6_FILE, TABLE6_ERROR_FILE,
    TABLE7_FILE, PHOTO_LEDGER_FILE, HEADER_ROW, parse_filename, parse_people,
)

# ============ 配置（与 main.py 保持同步） ============
THRESHOLD = 0.45
SAVE_ANNOTATED = False
REF_DIR = BASE_DIR / "Ref_Figure"
TARGET_DIR = BASE_DIR / "Target_Figure"
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
RESULT_FILE = OUTPUT_DIR / "00_程序人脸识别结果.xlsx"
FACE_DB_CACHE_FILE = REF_DIR / "face_db_cache.pkl"  # 随人脸库存放
# ====================================================

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".webp"}

# 输入表：web key → input/ 文件名
INPUT_FILES = {
    "photo_ledger": "01_照片台账表.xlsx",
    "work_plan": "02_工作安排表.xlsx",
    "person_class": "03_人员分类表.xlsx",
    "project_info": "04_项目信息表.xlsx",
    "attendance_template": "05_工程与外协考勤表模板.xlsx",
}

# 表格函数分发表
TABLE_FUNCS = {
    "table1": get_table1,
    "table2": get_table2,
    "table3": get_table3,
    "table4": get_table4,
    "table5": get_table5,
    "table6": get_table6,
    "table7": get_table7,
    "table8_9_10": get_table8_9_10,
}
TABLE_LABELS = {
    "table1": "表1 今日相机出工信息提取表",
    "table2": "表2 工作安排_提取表",
    "table3": "表3 出工照片人脸识别结果表",
    "table4": "表4 出工照片识别出工人表",
    "table5": "表5 该日出工人员表",
    "table6": "表6 全体人员出工情况表",
    "table7": "表7 出工地点及时长统计表",
    "table8_9_10": "表8/9/10 最终考勤表",
}
TABLE_OUTPUTS = {
    "table1": TABLE1_FILE,
    "table2": TABLE2_FILE,
    "table3": TABLE3_FILE,
    "table4": TABLE4_FILE,
    "table5": TABLE5_FILE,
    "table6": TABLE6_FILE,
    "table7": TABLE7_FILE,
}
# 最终考勤表：一个文件三个子表（表8/表9/表10）
COMBINED_FILE = "08_09_10_最终考勤表.xlsx"
FINAL_FILES = [COMBINED_FILE]

# 需要人工确认的表格（每次生成都弹窗）
CONFIRM_TABLES = {"table5": "05", "table6": "06"}

# 输出文件 → 显示名
OUTPUT_FILES_LABELS = {
    "00_程序人脸识别结果.xlsx": "人脸识别结果",
    TABLE1_FILE.name: "表1",
    TABLE2_FILE.name: "表2",
    TABLE3_FILE.name: "表3",
    TABLE4_FILE.name: "表4",
    TABLE5_FILE.name: "表5",
    TABLE6_FILE.name: "表6",
    TABLE5_ERROR_FILE.name: "核对信息错误文档1",
    TABLE6_ERROR_FILE.name: "核对信息错误文档2",
    TABLE7_FILE.name: "表7",
    COMBINED_FILE: "最终考勤表（表8/9/10）",
}

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100MB

# 防止识别与表格生成并发写 output/
RUN_LOCK = threading.Lock()


# ============================================================
# 工具函数
# ============================================================
def _image_files(d):
    if not d.is_dir():
        return []
    return [p for p in d.iterdir() if p.suffix.lower() in IMAGE_EXTS]


def _unique_path(d, name):
    """同名文件自动加序号：stem_1.ext。"""
    p = d / name
    if not p.exists():
        return p
    stem, suffix = Path(name).stem, Path(name).suffix
    i = 1
    while (d / f"{stem}_{i}{suffix}").exists():
        i += 1
    return d / f"{stem}_{i}{suffix}"


def _list_outputs():
    outputs = []
    for filename, label in OUTPUT_FILES_LABELS.items():
        path = OUTPUT_DIR / filename
        if path.exists():
            outputs.append({
                "name": filename,
                "label": label,
                "size": path.stat().st_size,
                "mtime": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            })
    return outputs


def _clear_dir(d):
    """删除目录内的所有文件与子目录（保留目录本身），返回删除项数。"""
    if not d.is_dir():
        return 0
    n = 0
    for item in d.iterdir():
        try:
            if item.is_dir():
                import shutil

                shutil.rmtree(item)
            else:
                item.unlink()
            n += 1
        except Exception:
            pass
    return n


def _resolve_providers():
    """按可用性选择推理后端：CUDA → DirectML → CPU（本机无 CUDA 时回退 CPU）。"""
    try:
        import onnxruntime as ort
        avail = set(ort.get_available_providers())
        for p in ("CUDAExecutionProvider", "DmlExecutionProvider", "CPUExecutionProvider"):
            if p in avail:
                return [p, "CPUExecutionProvider"] if p != "CPUExecutionProvider" else [p]
    except Exception:
        pass
    return ["CPUExecutionProvider"]


# ============================================================
# 共享人脸模型（识别与人工复核共用同一实例，避免重复加载 ~1GB 模型）
# ============================================================
_FACE_LOCK = threading.Lock()
_face_app = None


def _get_face_app():
    """延迟加载并缓存 InsightFace 模型实例（进程内只加载一次）。"""
    global _face_app
    with _FACE_LOCK:
        if _face_app is None:
            from face_app import init_face_app

            _face_app = init_face_app(_resolve_providers())
    return _face_app


def _get_face_db(app_):
    """返回当前人脸库向量（命中缓存则直接复用，无需重新推理）。"""
    from face_app import enroll_faces

    return enroll_faces(app_, REF_DIR, False, None, FACE_DB_CACHE_FILE)


def _run_func(func):
    """直接调用表格函数，捕获 stdout 与返回值。返回 (success, output, result)。"""
    buf = io.StringIO()
    try:
        with contextlib.redirect_stdout(buf):
            result = func()
        return True, buf.getvalue(), result
    except Exception:
        return False, buf.getvalue() + "\n" + traceback.format_exc(), None


# ============================================================
# 人脸识别后台线程（实时日志）
# ============================================================
class _LiveStream(io.StringIO):
    """把 stdout 按行实时转发给回调，同时保留完整文本。"""

    def __init__(self, on_line):
        super().__init__()
        self._on_line = on_line
        self._buf = ""

    def write(self, s):
        super().write(s)
        self._buf += s
        while "\n" in self._buf:
            line, self._buf = self._buf.split("\n", 1)
            if line:
                self._on_line(line)
        return len(s)

    def drain(self):
        if self._buf:
            self._on_line(self._buf)
            self._buf = ""


class RecognitionRunner:
    """后台运行人脸识别，前端轮询 /api/progress 获取实时日志。"""

    def __init__(self):
        self._lock = threading.Lock()
        self.running = False
        self.done = False
        self.success = False
        self.message = ""
        self.log_lines = []
        self.result_file = None

    def snapshot(self):
        with self._lock:
            return {
                "running": self.running,
                "done": self.done,
                "success": self.success,
                "message": self.message,
                "logs": list(self.log_lines),
                "result_file": str(self.result_file) if self.result_file else None,
            }

    def reset(self):
        """清空识别状态（用于清空文件夹后，让界面回到未识别状态）。"""
        with self._lock:
            self.done = False
            self.success = False
            self.message = ""
            self.log_lines = []
            self.result_file = None

    def start(self):
        if self.running:
            return False
        with self._lock:
            self.done = False
            self.success = False
            self.message = ""
            self.log_lines = []
            self.result_file = None
        RUN_LOCK.acquire()  # 整个识别期间独占，防止与表格生成冲突
        with self._lock:
            self.running = True
        threading.Thread(target=self._run, daemon=True).start()
        return True

    def _run(self):
        stream = _LiveStream(lambda line: self._add_log([line]))
        try:
            with contextlib.redirect_stdout(stream):
                # 惰性导入：仅此线程需要 insightface / cv2
                from face_app import (
                    enroll_faces, recognize_faces, save_results_to_excel,
                )

                app_ = _get_face_app()  # 与人工复核共享同一模型实例
                face_db = enroll_faces(app_, REF_DIR, SAVE_ANNOTATED, OUTPUT_DIR, FACE_DB_CACHE_FILE)
                results = recognize_faces(
                    app_, face_db, TARGET_DIR, THRESHOLD, SAVE_ANNOTATED, OUTPUT_DIR
                )
                if results:
                    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
                    save_results_to_excel(results, RESULT_FILE)
                    self.result_file = RESULT_FILE
                    self.message = f"识别完成，共处理 {len(results)} 张照片"
                    review_runner.invalidate()  # 新识别结果使旧复核缓存失效
                else:
                    self.message = "未检测到可识别照片（Target_Figure 中可能没有检测到人脸）"
            self.success = True
        except ModuleNotFoundError:
            self.message = (
                "未安装 insightface / opencv，无法执行人脸识别。\n"
                "请使用 AI_PY312 环境启动本程序（C:/Users/29037/.conda/envs/AI_PY312/python.exe web_show/web_app.py）。"
            )
            self.success = False
        except Exception:
            self.message = traceback.format_exc()
            self.success = False
        finally:
            stream.drain()
            with self._lock:
                self.running = False
                self.done = True
            RUN_LOCK.release()

    def _add_log(self, lines):
        with self._lock:
            self.log_lines.extend(lines)


runner = RecognitionRunner()


# ============================================================
# 人工复核数据后台构建（模型加载 + 逐张重检照片耗时数十秒，
# 不能在 HTTP 请求里同步阻塞，否则前端一直转圈）
# ============================================================
class ReviewBuildRunner:
    """后台构建人工复核数据，前端轮询 /api/review/progress 获取实时进度。"""

    def __init__(self):
        self._lock = threading.Lock()
        self.running = False
        self.done = False
        self.success = False
        self.message = ""
        self.progress = ""
        self.data = None
        self._cancel = False

    def snapshot(self):
        with self._lock:
            return {
                "running": self.running,
                "done": self.done,
                "success": self.success,
                "message": self.message,
                "progress": self.progress,
                "has_data": self.data is not None,
            }

    def invalidate(self):
        """识别结果被重新生成 / 文件夹被清空后，缓存失效，下次 openReview 重建。"""
        with self._lock:
            self.done = False
            self.success = False
            self.data = None
            self.progress = ""

    def get_data(self):
        with self._lock:
            return self.data

    def start(self):
        """启动构建。已在构建中或已有有效缓存时返回 False（不重复构建）。"""
        with self._lock:
            if self.running:
                return False
            if self.done and self.success and self.data is not None:
                return False
            self._cancel = False
            self.running = True
            self.done = False
            self.success = False
            self.message = ""
            self.progress = "正在准备…"
            self.data = None
        threading.Thread(target=self._run, daemon=True).start()
        return True

    def cancel(self):
        """请求取消构建。构建线程会在下一张照片前检查标志并释放锁。"""
        with self._lock:
            self._cancel = True

    def _is_cancelled(self):
        with self._lock:
            return self._cancel

    def _set_progress(self, text):
        with self._lock:
            self.progress = text

    def _run(self):
        try:
            RUN_LOCK.acquire()  # 与识别/表格生成互斥，避免并发使用模型
            try:
                data = _build_review_data(self._set_progress, self._is_cancelled)
            finally:
                RUN_LOCK.release()
            with self._lock:
                self.data = data
                self.success = True
                self.message = "完成"
        except ReviewCancelled:
            with self._lock:
                self.message = "已取消"
                self.success = False
                self.data = None
        except Exception:
            with self._lock:
                self.message = traceback.format_exc()
                self.success = False
        finally:
            with self._lock:
                self.running = False
                self.done = True

    def update_after_submit(self, photo):
        """写回 00_ 后就地移除该照片，保持缓存新鲜，避免再次重检。"""
        with self._lock:
            if self.data and self.data.get("photos") is not None:
                self.data["photos"] = [p for p in self.data["photos"] if p.get("photo") != photo]
                self.data["needs_review"] = len(self.data["photos"]) > 0


review_runner = ReviewBuildRunner()


# ============================================================
# 页面 / 静态资源
# ============================================================
@app.route("/")
def index():
    return send_file(WEB_SHOW_DIR / "index.html")


@app.route("/style.css")
def style_css():
    return send_file(WEB_SHOW_DIR / "style.css")


@app.route("/logo.png")
def logo_png():
    return send_file(WEB_SHOW_DIR / "logo.png", mimetype="image/png")


@app.route("/favicon.ico")
def favicon():
    return send_file(WEB_SHOW_DIR / "logo.png", mimetype="image/png")


@app.route("/app.js")
def app_js():
    return send_file(WEB_SHOW_DIR / "app.js")


# ============================================================
# 状态
# ============================================================
def _review_meta():
    """从 00_ 廉价统计复核信息（只读表格，不加载人脸模型）。

    返回 {"review_pending", "review_photos", "no_face_photos"}：
      review_pending  是否有照片含未识别(Unknown)的人脸，需要人工复核
      review_photos   含未识别人脸的照片数
      no_face_photos  完全未检测到人脸的照片数（不纳入复核，仅提示）
    """
    meta = {"review_pending": False, "review_photos": 0, "no_face_photos": 0}
    if not RESULT_FILE.exists():
        return meta
    try:
        df = pd.read_excel(RESULT_FILE)
        names = df["识别人名"].fillna("").astype(str)
        meta["review_photos"] = len(df.loc[names == "Unknown", "被识别图像名称"].unique())
        meta["no_face_photos"] = len(df.loc[names == "未检测到人脸", "被识别图像名称"].unique())
        meta["review_pending"] = meta["review_photos"] > 0
    except Exception:
        pass
    return meta


@app.route("/api/status")
def api_status():
    uploads = {key: (INPUT_DIR / fname).exists() for key, fname in INPUT_FILES.items()}
    snap = runner.snapshot()
    tables = {key: path.exists() for key, path in TABLE_OUTPUTS.items()}
    tables["table8_9_10"] = all((OUTPUT_DIR / n).exists() for n in FINAL_FILES)
    return jsonify({
        "uploads": uploads,
        "image_counts": {
            "ref": len(_image_files(REF_DIR)),
            "target": len(_image_files(TARGET_DIR)),
        },
        "outputs": _list_outputs(),
        "tables": tables,
        "recognition_running": snap["running"],
        "recognition_done": (snap["done"] and snap["success"]) or RESULT_FILE.exists(),
        **_review_meta(),
    })


# ============================================================
# 上传
# ============================================================
@app.route("/api/upload_images/<zone>", methods=["POST"])
def upload_images(zone):
    if zone not in ("ref", "target"):
        return jsonify({"success": False, "message": "未知上传区"}), 400
    target_dir = REF_DIR if zone == "ref" else TARGET_DIR
    target_dir.mkdir(parents=True, exist_ok=True)

    files = request.files.getlist("files")
    if not files:
        return jsonify({"success": False, "message": "未收到文件"}), 400

    saved, rejected = [], []
    for f in files:
        name = Path(f.filename or "").name
        if not name or "/" in name or "\\" in name or ".." in name:
            rejected.append(f.filename)
            continue
        if Path(name).suffix.lower() not in IMAGE_EXTS:
            rejected.append(name)
            continue
        try:
            f.save(_unique_path(target_dir, name))
            saved.append(name)
        except Exception as e:
            rejected.append(f"{name}({e})")

    return jsonify({
        "success": True,
        "saved": saved,
        "rejected": rejected,
        "image_counts": {
            "ref": len(_image_files(REF_DIR)),
            "target": len(_image_files(TARGET_DIR)),
        },
    })


@app.route("/api/upload_input/<file_key>", methods=["POST"])
def upload_input(file_key):
    if file_key not in INPUT_FILES:
        return jsonify({"success": False, "message": "未知文件类型"}), 400
    fname = INPUT_FILES[file_key]
    if "file" not in request.files:
        return jsonify({"success": False, "message": "未收到文件"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"success": False, "message": "文件名为空"}), 400
    if Path(f.filename).suffix.lower() != ".xlsx":
        return jsonify({"success": False, "message": "请输入 .xlsx 文件"}), 400
    save_path = INPUT_DIR / fname
    try:
        save_path.parent.mkdir(parents=True, exist_ok=True)
        f.save(save_path)
        return jsonify({"success": True, "message": f"已保存：input/{fname}"})
    except Exception as e:
        return jsonify({"success": False, "message": f"保存失败: {e}"}), 500


# ============================================================
# 人脸识别
# ============================================================
@app.route("/api/recognize", methods=["POST"])
def recognize():
    snap = runner.snapshot()
    if snap["running"]:
        return jsonify({
            "success": False, "started": False, "running": True,
            "message": "人脸识别已在进行中，请稍候",
        }), 409
    if review_runner.snapshot()["running"]:
        # 复核数据正在后台准备：取消它并等它释放模型锁，然后照常开始识别
        review_runner.cancel()
        for _ in range(20):  # 最多等约 10 秒
            if not review_runner.snapshot()["running"]:
                break
            time.sleep(0.5)
        if review_runner.snapshot()["running"]:
            return jsonify({
                "success": False, "started": False, "running": False,
                "message": "人工复核数据准备未能及时停止，请稍候再试",
            }), 409
    if not _image_files(REF_DIR):
        return jsonify({
            "success": False, "started": False, "running": False,
            "message": "Ref_Figure（员工人脸库）内没有图片，请先上传",
        }), 400
    if not _image_files(TARGET_DIR):
        return jsonify({
            "success": False, "started": False, "running": False,
            "message": "Target_Figure（出工照片）内没有图片，请先上传",
        }), 400
    started = runner.start()
    return jsonify({
        "success": started,
        "started": started,
        "running": started,
        "message": "人脸识别已启动" if started else "启动失败",
    })


@app.route("/api/progress")
def progress():
    return jsonify(runner.snapshot())


# ============================================================
# 人工复核（对 00_ 中未识别的人脸逐张人工修正）
# ============================================================
def _normalize_face_seq(v):
    """人脸列归一化：NaN/None/空 → ""，其余转字符串。"""
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    return str(v).strip()


def _norm_time(v) -> str:
    """时间值归一化为 'YYYY-MM-DD HH:MM:SS' 字符串，用于照片与台账匹配。"""
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    try:
        return pd.to_datetime(v).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return str(v).strip()


class ReviewCancelled(Exception):
    """复核数据构建被用户取消（例如用户改点「开始人脸识别」）。"""


def _build_review_data(progress_cb=None, cancel_cb=None):
    """
    构建人工复核数据：读取 00_，对含 Unknown 人脸的照片重新检测，
    给出未识别人脸的框坐标、上报/识别信息与差异提示。

    progress_cb(text)：可选进度回调，用于在后台线程里向界面实时报告进度。
    cancel_cb()：可选取消检查，返回 True 时抛出 ReviewCancelled 中止构建。
    """
    if cancel_cb and cancel_cb():
        raise ReviewCancelled()
    if progress_cb:
        progress_cb("正在读取识别结果…")
    if not RESULT_FILE.exists():
        return {"needs_review": False, "message": "尚未生成 00_程序人脸识别结果.xlsx"}
    df = pd.read_excel(RESULT_FILE)
    if not {"被识别图像名称", "识别人名"}.issubset(df.columns):
        return {"needs_review": False, "message": "00_ 识别结果格式异常（缺少必要列）"}

    df["被识别图像名称"] = df["被识别图像名称"].fillna("").astype(str)
    df["识别人名"] = df["识别人名"].fillna("").astype(str)
    df["人脸"] = df["人脸"].map(_normalize_face_seq)

    no_face_photos = df.loc[df["识别人名"] == "未检测到人脸", "被识别图像名称"].unique()
    pending_photos = sorted(df.loc[df["识别人名"] == "Unknown", "被识别图像名称"].unique())

    # 读取照片台账表，建立 (拍摄人, 拍摄时间) → 人员名单 索引（用于显示照片人员名单）
    if progress_cb:
        progress_cb("正在读取照片台账表…")
    ledger_index = {}
    ledger_loaded = False
    try:
        if PHOTO_LEDGER_FILE.exists():
            ldf = pd.read_excel(PHOTO_LEDGER_FILE, header=HEADER_ROW)
            for _, lr in ldf.iterrows():
                ph = str(lr.get("拍摄人", "")).strip() if pd.notna(lr.get("拍摄人")) else ""
                key = (ph, _norm_time(lr.get("拍摄时间")))
                ledger_index[key] = parse_people(lr.get("人员名单"))
            ledger_loaded = True
    except Exception:
        ledger_index = {}
        ledger_loaded = False

    # 惰性加载模型 + 人脸库（复用缓存向量）
    from face_app.face_engine import match_face
    from face_app.utils import imread_unicode

    if progress_cb:
        progress_cb("正在加载人脸模型…")
    app_ = _get_face_app()
    if progress_cb:
        progress_cb("正在读取人脸库…")
    face_db = _get_face_db(app_)
    known_names = sorted(face_db.keys())

    photos = []
    total = len(pending_photos)
    for idx, photo in enumerate(pending_photos, 1):
        if cancel_cb and cancel_cb():
            raise ReviewCancelled()
        if progress_cb:
            progress_cb(f"正在检测照片 {idx}/{total}…")
        rows = df[df["被识别图像名称"] == photo].reset_index(drop=True)
        reporter, photo_time, _ = parse_filename(photo)
        recognized_names = sorted({
            r for r in rows["识别人名"]
            if r and r not in ("Unknown", "未检测到人脸")
        })

        # 照片人员名单：优先取自照片台账表（按 拍摄人+拍摄时间 匹配）
        person_list = ledger_index.get((reporter, _norm_time(photo_time)), [])
        list_warning = ""
        if not ledger_loaded:
            list_warning = "照片台账表未上传，无法读取人员名单"
        elif not person_list:
            list_warning = "未在照片台账表匹配到该照片的人员名单"

        unknown_faces = []
        warning = ""
        img_path = TARGET_DIR / photo
        if not img_path.exists():
            warning = "原图不在 Target_Figure 中"
        else:
            img = imread_unicode(str(img_path))
            if img is None:
                warning = "无法读取原图"
            else:
                faces = app_.get(img)
                if len(faces) != len(rows):
                    warning = f"重新检测出 {len(faces)} 张脸，与识别记录 {len(rows)} 条不一致"
                else:
                    for i, face in enumerate(faces):
                        name, score = match_face(face.normed_embedding, face_db, THRESHOLD)
                        if name != "Unknown":
                            continue
                        unknown_faces.append({
                            "seq": rows.at[i, "人脸"],  # 与 00_ 行对应，用于写回
                            "box": [int(v) for v in face.bbox],
                            "score": round(float(score), 4),  # float32 → Python float，否则 JSON 序列化失败
                        })

        # 差异提示（拍照人无法上镜，出现在名单但未被识别不算异常，故排除）
        diff_report_not_recognized = [
            n for n in person_list
            if n not in recognized_names and n != reporter
        ]
        diff_recognized_not_report = [n for n in recognized_names if n not in person_list]

        photos.append({
            "photo": photo,
            "reporter": reporter,           # 拍照人（用于排除判断，不判异常）
            "person_list": person_list,     # 照片人员名单（来自台账表）
            "list_warning": list_warning,
            "recognized_names": recognized_names,
            "unknown_faces": unknown_faces,
            "warning": warning,
            "diffs": {
                "report_not_recognized": diff_report_not_recognized,
                "recognized_not_report": diff_recognized_not_report,
            },
        })

    return {
        "needs_review": len(photos) > 0,
        "photos": photos,
        "no_face_photos": len(no_face_photos),
        "known_names": known_names,
    }


@app.route("/api/review/start", methods=["POST"])
def review_start():
    """启动后台构建复核数据；已在构建或已有有效缓存时直接返回当前状态。"""
    if runner.snapshot()["running"]:
        return jsonify({"success": False, "message": "人脸识别运行中，请稍候"}), 409
    review_runner.start()
    # 注意：snapshot() 里的 success 表示"构建是否完成"，放在前面，不能覆盖 API 的 success
    return jsonify({**review_runner.snapshot(), "success": True})


@app.route("/api/review/progress")
def review_progress():
    return jsonify({**review_runner.snapshot()})


@app.route("/api/review/data")
def review_data():
    snap = review_runner.snapshot()
    if snap["running"]:
        return jsonify({"success": False, "message": "正在准备中"}), 425
    if not snap["done"] or not snap["success"] or not snap["has_data"]:
        return jsonify({"success": False, "message": snap.get("message") or "暂无复核数据"}), 425
    data = review_runner.get_data()
    return jsonify({"success": True, **data})


@app.route("/api/review/cancel", methods=["POST"])
def review_cancel():
    """取消正在进行的复核数据构建（例如用户改点「开始人脸识别」/「跳过本次复核」）。"""
    review_runner.cancel()
    # 同上：success 放在最后，表示 API 本身成功
    return jsonify({**review_runner.snapshot(), "success": True})


@app.route("/api/review_submit", methods=["POST"])
def review_submit():
    if runner.snapshot()["running"]:
        return jsonify({"success": False, "message": "人脸识别运行中，请稍候"}), 409
    body = request.get_json(silent=True) or {}
    photo = str(body.get("photo", "")).strip()
    corrections = body.get("corrections") or []
    if not photo or not RESULT_FILE.exists():
        return jsonify({"success": False, "message": "参数缺失或 00_ 不存在"}), 400

    cleaned = []
    for c in corrections:
        seq = _normalize_face_seq(c.get("seq"))
        name = str(c.get("name", "")).strip()
        if name:
            cleaned.append({"seq": seq, "name": name})
    if not cleaned:
        return jsonify({"success": True, "updated": 0, "message": "未填写任何姓名"})

    with RUN_LOCK:
        try:
            df = pd.read_excel(RESULT_FILE)
            if "修正来源" not in df.columns:
                df["修正来源"] = ""
            df["被识别图像名称"] = df["被识别图像名称"].fillna("").astype(str)
            df["识别人名"] = df["识别人名"].fillna("").astype(str)
            df["人脸"] = df["人脸"].map(_normalize_face_seq)

            updated = 0
            for c in cleaned:
                mask = (df["被识别图像名称"] == photo) & (df["人脸"] == c["seq"])
                if mask.any():
                    df.loc[mask, "识别人名"] = c["name"]
                    df.loc[mask, "状态"] = "已识别"
                    df.loc[mask, "修正来源"] = "人工"
                    updated += int(mask.sum())
            if updated > 0:  # 只有实际改到行才写回文件，避免空写
                df.to_excel(RESULT_FILE, index=False)
        except Exception:
            return jsonify({
                "success": False,
                "message": f"写回 00_ 失败：\n{traceback.format_exc()}",
            }), 500

    # 缓存里的复核数据就地移除该照片，保持新鲜，无需重新检测
    review_runner.update_after_submit(photo)

    return jsonify({
        "success": True,
        "updated": updated,
        "message": f"已更新 {updated} 条识别记录" if updated else "未填写姓名，已跳过本张",
    })


@app.route("/api/target_image/<name>")
def target_image(name):
    """服务 Target_Figure 原图（人工复核照片展示用）。"""
    safe = Path(name).name  # 防路径逃逸
    if not (TARGET_DIR / safe).exists():
        return jsonify({"success": False, "message": f"图片不存在: {safe}"}), 404
    return send_from_directory(str(TARGET_DIR), safe)


# ============================================================
# 表格生成
# ============================================================
@app.route("/api/run_table/<key>", methods=["POST"])
def run_table(key):
    if key not in TABLE_FUNCS:
        return jsonify({"success": False, "message": f"未知表格: {key}"}), 404
    if runner.snapshot()["running"]:
        return jsonify({"success": False, "message": "人脸识别正在运行，请稍后再试"}), 409

    with RUN_LOCK:
        success, output, result = _run_func(TABLE_FUNCS[key])

    info = None
    if key in CONFIRM_TABLES and success and isinstance(result, dict):
        info = {
            "normal": result.get("normal"),
            "abnormal": result.get("abnormal"),
            "error_doc": Path(result["error_doc"]).name if result.get("error_doc") else None,
            "table_file": Path(result["table_file"]).name if result.get("table_file") else None,
        }

    return jsonify({
        "success": success,
        "key": key,
        "label": TABLE_LABELS.get(key, key),
        "message": "执行成功" if success else "执行失败，请查看日志",
        "output": output,
        "confirm_required": key in CONFIRM_TABLES and success,
        "table_info": info,
        "outputs": _list_outputs(),
    })


@app.route("/api/confirm/<target>", methods=["POST"])
def confirm(target):
    mapping = {"05": TABLE5_FILE, "06": TABLE6_FILE}
    if target not in mapping:
        return jsonify({"success": False, "message": "未知确认目标"}), 400
    dest = mapping[target]
    f = request.files.get("file")
    if f and f.filename:
        if Path(f.filename).suffix.lower() != ".xlsx":
            return jsonify({"success": False, "message": "回传修正表必须是 .xlsx 文件"}), 400
        try:
            dest.parent.mkdir(parents=True, exist_ok=True)
            f.save(dest)
        except Exception as e:
            return jsonify({"success": False, "message": f"保存失败: {e}"}), 500
    return jsonify({
        "success": True,
        "message": "已确认，继续",
        "outputs": _list_outputs(),
    })


# ============================================================
# 表格展示 / 在线编辑（表1、核对信息错误文档、最终表、05/06）
# ============================================================
def _esc_html(v) -> str:
    """转义单元格值用于 HTML 输出。"""
    if v is None:
        return ""
    s = str(v)
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def sheet_to_html(ws, highlight_cols=None) -> str:
    """
    把 openpyxl 工作表渲染为 HTML <table>，保留合并单元格（colspan/rowspan）。
    highlight_cols: 列名列表，非空单元格加 class="hl-anomaly"。
    """
    highlight_cols = highlight_cols or []

    # 第一行（表头）建立 列名 → 列下标
    col_name_to_idx = {}
    for cell in ws[1]:
        if cell.value is not None:
            col_name_to_idx[str(cell.value).strip()] = cell.column - 1
    highlight_idxs = {
        col_name_to_idx[c] for c in highlight_cols if c in col_name_to_idx
    }

    # 合并范围 → 左上角坐标 → (colspan, rowspan)
    merge_map = {}
    for mr in ws.merged_cells.ranges:
        merge_map[(mr.min_row, mr.min_col)] = (
            mr.max_col - mr.min_col + 1,
            mr.max_row - mr.min_row + 1,
        )
    covered = set()
    for (r0, c0), (span_c, span_r) in merge_map.items():
        for r in range(r0, r0 + span_r):
            for c in range(c0, c0 + span_c):
                if (r, c) != (r0, c0):
                    covered.add((r, c))

    # 首列表头为"姓名"时固定该列（横向滚动保持不动）；错误文档等文本表不加
    first_cell = ws.cell(row=1, column=1).value
    sticky = (
        " sticky-name"
        if first_cell is not None and str(first_cell).strip() == "姓名"
        else ""
    )
    parts = [f'<table class="sheet-table{sticky}">']
    for row in ws.iter_rows():
        parts.append("<tr>")
        for cell in row:
            rc = (cell.row, cell.column)
            if rc in covered:
                continue
            attrs = ""
            if rc in merge_map:
                span_c, span_r = merge_map[rc]
                attrs += f' colspan="{span_c}" rowspan="{span_r}"'
            if cell.column - 1 in highlight_idxs and cell.value not in (None, ""):
                attrs += ' class="hl-anomaly"'
            parts.append(f"<td{attrs}>{_esc_html(cell.value)}</td>")
        parts.append("</tr>")
    parts.append("</table>")
    return "".join(parts)


# 可查看表：key → (文件名, 标题)
VIEW_TABLES = {
    "table1": ("01_今日相机出工信息提取表.xlsx", "表1 今日相机出工信息提取表"),
    "error1": ("核对信息错误文档1.xlsx", "核对信息错误文档1"),
    "error2": ("核对信息错误文档2.xlsx", "核对信息错误文档2"),
    "final": (COMBINED_FILE, "最终考勤表（表8/9/10）"),
}
# 需要高亮异常提示的列
VIEW_HIGHLIGHT_COLS = {
    "table1": ["是否开工汇报异常", "是否收工汇报异常"],
}


@app.route("/api/view_sheet/<key>")
def view_sheet(key):
    if key not in VIEW_TABLES:
        return jsonify({"success": False, "message": "未知表格"}), 404
    fname, title = VIEW_TABLES[key]
    path = OUTPUT_DIR / fname
    if not path.exists():
        return jsonify({"success": False, "message": f"文件不存在: {fname}"}), 404
    try:
        wb = load_workbook(path, data_only=True)
        hc = VIEW_HIGHLIGHT_COLS.get(key, [])
        sheets = [
            {"name": ws.title, "html": sheet_to_html(ws, hc)}
            for ws in wb.worksheets
        ]
        wb.close()
    except Exception:
        return jsonify({
            "success": False,
            "message": f"读取失败: {fname}\n{traceback.format_exc()}",
        }), 500
    return jsonify({"success": True, "title": title, "sheets": sheets})


# 可编辑表：key → 文件 / 可编辑列 / 锁定列 / 待确认列
EDIT_TABLES = {
    "table5": {
        "file": TABLE5_FILE,
        "editable_cols": None,  # 全部可编辑
        "locked_cols": [],
        "pending_col": "加班时长确认-人工审核",
    },
    "table6": {
        "file": TABLE6_FILE,
        "editable_cols": ["开工项目名", "开工项目简称", "收工项目名", "收工项目简称"],
        "locked_cols": ["加班时长确认-人工审核"],
        "pending_col": "",
    },
}


@app.route("/api/edit_data/<key>")
def edit_data(key):
    if key not in EDIT_TABLES:
        return jsonify({"success": False, "message": "未知编辑表"}), 404
    info = EDIT_TABLES[key]
    path = info["file"]
    if not path.exists():
        return jsonify({"success": False, "message": f"文件不存在: {path.name}"}), 404
    df = pd.read_excel(path)
    columns = [str(c) for c in df.columns]
    rows = [
        ["" if pd.isna(v) else str(v) for v in r]
        for _, r in df.iterrows()
    ]
    editable = (
        columns if info.get("editable_cols") is None else info["editable_cols"]
    )
    return jsonify({
        "success": True,
        "columns": columns,
        "rows": rows,
        "editable_idxs": [columns.index(c) for c in editable if c in columns],
        "locked_idxs": [columns.index(c) for c in info["locked_cols"] if c in columns],
        "pending_idx": columns.index(info["pending_col"]) if info.get("pending_col") in columns else -1,
    })


def _apply_table6_highlights(path):
    """06 写回后重放高亮：空「开工/收工项目简称」、非空「备注」→ 黄色底。"""
    wb = load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]

    def col_idx(name):
        return header.index(name) + 1 if name in header else None

    kg = col_idx("开工项目简称")
    sg = col_idx("收工项目简称")
    rm = col_idx("备注")
    fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for r in range(2, ws.max_row + 1):
        for col, empty_expected in ((kg, True), (sg, True), (rm, False)):
            if col is None:
                continue
            val = ws.cell(row=r, column=col).value
            is_empty = val is None or str(val).strip() == ""
            if (empty_expected and is_empty) or (not empty_expected and not is_empty):
                ws.cell(row=r, column=col).fill = fill
    wb.save(path)
    wb.close()


@app.route("/api/edit_save/<key>", methods=["POST"])
def edit_save(key):
    if key not in EDIT_TABLES:
        return jsonify({"success": False, "message": "未知编辑表"}), 404
    info = EDIT_TABLES[key]
    path = info["file"]
    if not path.exists():
        return jsonify({"success": False, "message": f"文件不存在: {path.name}"}), 404
    body = request.get_json(silent=True) or {}
    columns = body.get("columns") or []
    rows = body.get("rows") or []
    if not columns or not isinstance(rows, list):
        return jsonify({"success": False, "message": "参数错误"}), 400

    # 以原始文件的列顺序为准，确定数值列
    old_df = pd.read_excel(path)
    old_cols = [str(c) for c in old_df.columns]
    numeric_cols = {
        str(c) for c in old_df.columns if pd.api.types.is_numeric_dtype(old_df[c])
    }
    if columns != old_cols:
        columns = old_cols
        rows = [
            [r[i] if i < len(r) else "" for i in range(len(old_cols))]
            for r in rows
        ]

    new_df = pd.DataFrame(rows, columns=columns)
    for c in numeric_cols:
        if c in new_df.columns:
            new_df[c] = pd.to_numeric(new_df[c], errors="coerce")

    # 锁定列强制恢复原值（如 06 的「加班时长确认-人工审核」只能在 05 修改，
    # 这里按姓名还原，防止网页端误改）
    locked_cols = [
        c for c in info.get("locked_cols", [])
        if c in old_df.columns and c in new_df.columns
    ]
    if locked_cols and "姓名" in old_df.columns and "姓名" in new_df.columns:
        for lc in locked_cols:
            orig_map = {}
            for n, v in zip(old_df["姓名"].astype(str).str.strip(), old_df[lc]):
                orig_map.setdefault(n, v)  # 首次出现为准
            new_df[lc] = new_df["姓名"].astype(str).str.strip().map(
                lambda n: orig_map.get(n, "")
            )

    new_df.to_excel(path, index=False)

    if key == "table6":
        _apply_table6_highlights(path)

    return jsonify({
        "success": True,
        "message": f"已保存 {len(new_df)} 行到 {path.name}",
    })


# ============================================================
# 清空文件夹（只清 output/ 和 Target_Figure/）
# ============================================================
@app.route("/api/clear_folders", methods=["POST"])
def clear_folders():
    if runner.snapshot()["running"]:
        return jsonify({
            "success": False,
            "message": "人脸识别正在运行，请先停止后再清空",
        }), 409
    with RUN_LOCK:
        n_out = _clear_dir(OUTPUT_DIR)
        n_tgt = _clear_dir(TARGET_DIR)
    runner.reset()
    review_runner.invalidate()  # 00_ 已删除，复核缓存失效
    return jsonify({
        "success": True,
        "message": f"已清空：output/（{n_out} 项）、Target_Figure/（{n_tgt} 项）。人脸库与输入表不受影响。",
        "outputs": _list_outputs(),
    })


# ============================================================
# 下载
# ============================================================
@app.route("/api/download/<path:filename>")
def download(filename):
    name = Path(filename).name  # 防止路径逃逸
    if not (OUTPUT_DIR / name).exists():
        return jsonify({"success": False, "message": f"文件不存在: {name}"}), 404
    return send_from_directory(str(OUTPUT_DIR), name, as_attachment=True)


# ============================================================
# 错误处理
# ============================================================
@app.errorhandler(413)
def too_large(_e):
    return jsonify({"success": False, "message": "上传文件过大（超过 100MB）"}), 413


if __name__ == "__main__":
    print("=" * 50)
    print(" 人脸识别考勤系统 Web 界面")
    print(" 浏览器打开: http://localhost:5000")
    print("=" * 50)
    app.run(host="0.0.0.0", port=5000, threaded=True, debug=False)
