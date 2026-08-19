"""
表生成模块 - 汇总表1~表10 的生成逻辑

每个 get_tableN() 对应原 generate_tableN.py 的 main()：

    get_table1()
    输入: input/01_照片台账表.xlsx → 输出: output/01_今日相机出工信息提取表.xlsx

    get_table2()
    输入: input/02_工作安排表.xlsx → 输出: output/02_工作安排_提取表.xlsx

    get_table3()
    输入: output/00_程序人脸识别结果.xlsx → 输出: ouput/03_出工照片人脸识别结果表.xlsx

    get_table4()
    输入: ouput/03_出工照片人脸识别结果表.xlsx → 输出: ouput/04_出工照片识别出工人表.xlsx

    get_table5()
    输入: output/01_今日相机出工信息提取表.xlsx + ouput/04_出工照片识别出工人表.xlsx
    输出: output/05_该日出工人员表.xlsx 或 核对信息错误文档1.xlsx
    Tips: 如果对比结果没有问题，直接输出 05_该日出工人员表.xlsx;
          如果对比结果有问题，输出“核对信息错误文档1.xlsx” 和 “05_该日出工人员表.xlsx”
          带核对信息完成之后自动将上传文件命名为"05_该日出工人员表.xlsx"放到output文件夹里面


    get_table6()
    输入: output/05_该日出工人员表.xlsx+ output/02_工作安排_提取表.xlsx + output/000_项目信息表.xlsx
    输出:output/06_全体人员出工情况表.xlsx或 核对信息错误文档2.xlsx
    Tips: 如果对比结果没有问题，直接输出 06_全体人员出工情况表.xlsx;
              如果对比结果有问题，输出“核对信息错误文档2.xlsx” 和 “06_全体人员出工情况表.xlsx”
              带核对信息完成之后自动将上传文件命名为"06_全体人员出工情况表.xlsx"放到output文件夹里面

    get_table7()
    输入: output/06_全体人员出工情况表.xlsx + input/03_人员分类表.xlsx
    输出: output/07_表7出工地点及时长统计表.xlsx

    get_table8_9_10()
    输入: output/07_表7出工地点及时长统计表.xlsx + input/05_工程与外协考勤表模板
    输出: output/08_09_10_最终考勤表.xlsx（一个文件三个子表：表8/表9/表10）

"""

import re
import difflib
from pathlib import Path
from copy import copy
from datetime import time

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 路径与配置
# ============================================================
BASE_DIR = Path(__file__).resolve().parent.parent
INPUT_DIR = BASE_DIR / "input"  # 输入表目录
OUTPUT_DIR = BASE_DIR / "output"  # 输出表 / 识别结果 / 缓存统一目录
INPUT_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# --- 表1 ---
PHOTO_LEDGER_FILE = INPUT_DIR / "01_照片台账表.xlsx"
TABLE1_FILE = OUTPUT_DIR / "01_今日相机出工信息提取表.xlsx"
HEADER_ROW = 2  # 照片台账表表头所在行（从 0 开始计数，第 3 行）
LATE_START_THRESHOLD = time(10, 0, 0)  # 开工晚于 10:00 记为开工晚发
OVERTIME_BASE_HOUR = 17  # 17:00 后开始计算加班

# --- 表2 ---
WORK_PLAN_FILE = INPUT_DIR / "02_工作安排表.xlsx"
TABLE2_FILE = OUTPUT_DIR / "02_工作安排_提取表.xlsx"
HEADER_ROW2 = 1  # 工作安排表表头所在行（第 2 行）
STATUS_ROWS = {
    "公司请假人员": "请假",
    "协助未到岗人员": "外协未到岗",
    "安排休息人员": "安排休息",
}

# --- 表3 ---
ATTENDANCE_FILE = OUTPUT_DIR / "00_程序人脸识别结果.xlsx"
TABLE3_FILE = OUTPUT_DIR / "03_出工照片人脸识别结果表.xlsx"
# 人工复核时确认"照片确实无人脸"写入 00_ 的标记值（与 Unknown/未检测到人脸 一样不当作出工人）
NO_FACE_CONFIRMED = "无人脸(人工确认)"

# --- 表4 ---
TABLE4_FILE = OUTPUT_DIR / "04_出工照片识别出工人表.xlsx"

# --- 表5（对比表1与表4，生成 05_该日出工人员表） ---
TABLE5_FILE = OUTPUT_DIR / "05_该日出工人员表.xlsx"
TABLE5_ERROR_FILE = OUTPUT_DIR / "核对信息错误文档1.xlsx"

# --- 表6（对比 05_ 与表2，生成 06_全体人员出工情况表） ---
TABLE6_FILE = OUTPUT_DIR / "06_全体人员出工情况表.xlsx"
TABLE6_ERROR_FILE = OUTPUT_DIR / "核对信息错误文档2.xlsx"
PROJECT_INFO_FILE = OUTPUT_DIR / "000_项目信息表.xlsx"  # 用于出勤统计简称匹配（由 project_info 生成）

# --- 表7 ---
CLASSIFY_FILE = INPUT_DIR / "03_人员分类表.xlsx"
TABLE7_FILE = OUTPUT_DIR / "07_表7出工地点及时长统计表.xlsx"
# 表6 中的名 → 001 中的标准名（OCR 字形误判纠错）
TABLE7_NAME_VARIANT_MAP = {
    "蔣连东": "蒋连东",  # 繁→简
    "陈题干": "陈提干",  # 题→提
}

# --- 表8/9/10 ---
TEMPLATE_FILE = INPUT_DIR / "05_工程与外协考勤表模板.xlsx"
# 表7 OCR 名 → 目标表标准名（在目标表中查找时纠错）
TABLE8_NAME_VARIANT_MAP = {
    "蔣连东": "蒋连东",  # 繁→简
    "陈题干": "陈提干",  # 题→提
    "吕洁": "吕杰",  # 洁→杰
}


# ============================================================
# 通用工具函数
# ============================================================
def parse_people(text: str) -> list[str]:
    """
    从【人员名单】单元格文本中解析出实际出工人名列表（表1/表2 共用）。

    规则：
    - 普通中文姓名（2~4 字）直接提取；
    - "包工头N人（实际名）"：括号内的人才是实际出工者，包工头本人不算；
    - "包工头N人"（无括号）：包工头本人出工，提取包工头。

    支持中文逗号、顿号、空格、英文逗号等分隔符。
    """
    if pd.isna(text):
        return []

    text = str(text).strip()

    # 先处理 "xxxN人（实际名）" 模式，提取括号内所有实际出工人名
    # 例: "张春林1人（张春均）"             → 替换为 "张春均"
    # 例: "张春林3人（张春林、蔡平波、蒋连东）" → 替换为 "张春林 蔡平波 蒋连东"
    # 例: "单工1人（许士祥）"                → 替换为 "许士祥"
    text = re.sub(
        r"[一-龥]+\d*人[（(]([^）)]+)[）)]",
        lambda m: " ".join(re.findall(r"[一-龥]{2,4}", m.group(1))),
        text,
    )

    # 再处理 "类别（实际名）" 模式（无"N人"数字连接，如 临时工（岳磊、梁建普））
    # 括号里的才是出工人员，类别本身不算
    text = re.sub(
        r"[一-龥]{2,6}[（(]([^）)]+)[）)]",
        lambda m: " ".join(re.findall(r"[一-龥]{2,4}", m.group(1))),
        text,
    )

    # 统一分隔符为空格
    text = re.sub(r"[、，,；;。.\\s]+", " ", text)

    people = []
    for token in text.split():
        # 提取 2~4 个连续汉字
        matches = re.findall(r"[一-龥]{2,4}", token)
        people.extend(matches)

    # 去重并保持顺序
    return list(dict.fromkeys(people))


def calc_overtime(off_time) -> float | None:
    """
    根据收工汇报时间计算加班时长（小时）。

    规则：
    - 收工时间 <= 17:00，加班 0 小时；
    - 17:00 之后，50 分钟以内计 0 小时，50~60 分钟计 1 小时；
    - 超过 1 小时后，整小时部分直接计入；
    - 余数 >= 50 分钟进 1 小时，>= 20 分钟计 0.5 小时，< 20 分钟舍去。

    示例：18:23 -> 1.5h；18:19 -> 1h；17:52 -> 1h；17:48 -> 0h。
    """
    if pd.isna(off_time):
        return None

    t = pd.to_datetime(off_time)
    base = t.replace(hour=OVERTIME_BASE_HOUR, minute=0, second=0, microsecond=0)
    if t <= base:
        return 0.0

    delta_minutes = (t - base).total_seconds() / 60

    # 不足 1 小时
    if delta_minutes < 50:
        return 0.0
    if delta_minutes < 60:
        return 1.0

    # 1 小时以上的部分
    hours = int(delta_minutes // 60)
    remainder = delta_minutes % 60
    if remainder >= 50:
        return hours + 1.0
    elif remainder >= 20:
        return hours + 0.5
    else:
        return float(hours)


def format_datetime(value) -> str:
    """将时间值格式化为字符串，空值返回空字符串。"""
    if pd.isna(value):
        return ""
    return pd.to_datetime(value).strftime("%Y-%m-%d %H:%M:%S")


def join_unique(values, sep="、") -> str:
    """将可迭代对象去重后连接成字符串，忽略空值/NaN。"""
    cleaned = []
    for v in values:
        if pd.isna(v):
            continue
        s = str(v).strip()
        if s and s not in cleaned:
            cleaned.append(s)
    return sep.join(cleaned)


def parse_filename(filename: str) -> tuple[str, str, str]:
    """
    从图片文件名中解析：拍摄人姓名、拍摄时间、人脸序列号。

    文件名格式：
      "储少球 2026_06_08 09_05_17 开(收)工汇报.jpg"

    返回 (姓名, 时间字符串, 序列号)。
    注：人脸序号在结果表中单独一列（"人脸"列），此处始终为空。
    """
    name = ""
    time_str = ""
    seq = ""

    # 提取时间：YYYY_MM_DD HH_MM_SS
    time_match = re.search(
        r"(\d{4})_(\d{2})_(\d{2})\s+(\d{2})_(\d{2})_(\d{2})", filename
    )
    if time_match:
        time_str = (
            f"{time_match.group(1)}-{time_match.group(2)}-{time_match.group(3)} "
            f"{time_match.group(4)}:{time_match.group(5)}:{time_match.group(6)}"
        )
        # 提取时间之前的文本作为人名
        before_time = filename[: time_match.start()].strip()
        name = before_time
    else:
        name = filename

    return name, time_str, seq


def collect_times(row, cols: list[str]) -> set[str]:
    """从一行中提取所有非空时间点（表5 用）。"""
    times = set()
    for col in cols:
        val = row.get(col, "")
        if pd.notna(val) and str(val).strip() != "":
            times.add(str(val).strip())
    return times


# ============================================================
# 表1
# ============================================================
def get_table1():
    """
    基于【照片台账表】生成【表1今日相机出工信息提取表】。

    按人汇总当天开工/收工信息（开工取最早、收工取最晚），
    判断"开工未发 / 开工晚发 / 收工未发 / 半天换场"，测算加班时长，
    对异常单元格高亮后输出 Excel。
    """
    print("=" * 50)
    print("       生成【表1今日相机出工信息提取表】")
    print("=" * 50)

    # 1. 读取照片台账表
    if not PHOTO_LEDGER_FILE.exists():
        raise FileNotFoundError(f"未找到照片台账表：{PHOTO_LEDGER_FILE}")

    df = pd.read_excel(PHOTO_LEDGER_FILE, header=HEADER_ROW)
    print(f"读取照片台账表：共 {len(df)} 条记录")

    # 2. 按人汇总照片记录
    person_records: dict[str, list[dict]] = {}

    for _, row in df.iterrows():
        photographer = row.get("拍摄人", "")
        people_text = row.get("人员名单", "")
        photo_time = row.get("拍摄时间", None)
        report_type = row.get("汇报类型", "")
        project = row.get("工程名称", "")
        location = row.get("拍摄地点", "")
        issue = row.get("存在问题", "")
        overtime_report = row.get("加班时长", None)

        # 构建该照片涉及的人员列表
        people = []
        if not pd.isna(photographer):
            people.append(str(photographer).strip())
        people.extend(parse_people(people_text))

        # 去重
        seen = set()
        unique_people = []
        for name in people:
            if name and name not in seen:
                seen.add(name)
                unique_people.append(name)

        # 为每个人登记该条记录
        for person in unique_people:
            person_records.setdefault(person, []).append(
                {
                    "拍摄时间": photo_time,
                    "汇报类型": report_type,
                    "工程名称": project,
                    "拍摄人": (
                        str(photographer).strip() if not pd.isna(photographer) else ""
                    ),
                    "拍摄地点": location,
                    "存在问题": issue,
                    "加班时长": overtime_report,
                }
            )

    # 3. 生成表1
    rows = []
    for person in sorted(person_records.keys()):
        records = person_records[person]

        start_records = [r for r in records if r["汇报类型"] == "开工汇报"]
        end_records = [r for r in records if r["汇报类型"] == "收工汇报"]

        # 开工：取最早时间
        start_time = None
        start_project = ""
        if start_records:
            start_time = min(r["拍摄时间"] for r in start_records)
            for r in start_records:
                if r["拍摄时间"] == start_time:
                    start_project = r["工程名称"]
                    break

        # 收工：取最晚时间
        end_time = None
        end_project = ""
        if end_records:
            end_time = max(r["拍摄时间"] for r in end_records)
            for r in end_records:
                if r["拍摄时间"] == end_time:
                    end_project = r["工程名称"]
                    break

        # 异常判断
        start_abnormal = ""
        if start_time is None:
            start_abnormal = "开工未发"
        elif pd.to_datetime(start_time).time() > LATE_START_THRESHOLD:
            start_abnormal = "开工晚发"

        end_abnormal = "收工未发" if end_time is None else ""

        # 加班时长
        overtime_calc = calc_overtime(end_time)
        overtime_report_values = [
            r["加班时长"] for r in records if not pd.isna(r["加班时长"])
        ]
        overtime_report = overtime_report_values[0] if overtime_report_values else ""

        # 汇总拍摄人（开工/收工分开）、地点、存在问题
        start_photographers = [r["拍摄人"] for r in start_records]
        end_photographers = [r["拍摄人"] for r in end_records]
        locations = [r["拍摄地点"] for r in records]
        issues = [r["存在问题"] for r in records]

        # 全天状态
        if start_project and end_project:
            all_day_status = "全天一致" if start_project == end_project else "半天换场"
        else:
            all_day_status = ""

        rows.append(
            {
                "姓名": person,
                "开工汇报拍摄时间": format_datetime(start_time),
                "是否开工汇报异常": start_abnormal,
                "开工项目名": start_project,
                "开工拍摄人": join_unique(start_photographers),
                "收工汇报拍摄时间": format_datetime(end_time),
                "是否收工汇报异常": end_abnormal,
                "收工项目名": end_project,
                "收工拍摄人": join_unique(end_photographers),
                "加班时长-基于收工时间测算": (
                    overtime_calc if overtime_calc is not None else ""
                ),
                "加班时长-上报时长": overtime_report,
                "加班时长确认-人工审核": "",
                "拍摄地点": join_unique(locations),
                "存在问题": join_unique(issues),
                "全天状态": all_day_status,
            }
        )

    # 4. 输出 Excel（含高亮标识）
    columns = [
        "姓名",
        "开工汇报拍摄时间",
        "是否开工汇报异常",
        "开工项目名",
        "开工拍摄人",
        "收工汇报拍摄时间",
        "是否收工汇报异常",
        "收工项目名",
        "收工拍摄人",
        "加班时长-基于收工时间测算",
        "加班时长-上报时长",
        "加班时长确认-人工审核",
        "拍摄地点",
        "存在问题",
        "全天状态",
    ]
    result_df = pd.DataFrame(rows, columns=columns)
    result_df.to_excel(TABLE1_FILE, index=False)

    # 高亮标识（半天换场 / 加班时长非0 / 上报时长非空）
    wb = load_workbook(TABLE1_FILE)
    ws = wb.active
    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    orange_fill = PatternFill(
        start_color="FFD700", end_color="FFD700", fill_type="solid"
    )
    green_fill = PatternFill(
        start_color="90EE90", end_color="90EE90", fill_type="solid"
    )

    status_col_idx = columns.index("全天状态") + 1
    overtime_calc_col_idx = columns.index("加班时长-基于收工时间测算") + 1
    overtime_report_col_idx = columns.index("加班时长-上报时长") + 1

    for row_idx in range(2, ws.max_row + 1):  # 第 1 行是表头
        # 半天换场 → 黄色
        if ws.cell(row=row_idx, column=status_col_idx).value == "半天换场":
            ws.cell(row=row_idx, column=status_col_idx).fill = yellow_fill

        # 加班时长-基于收工时间测算 非0 → 橙色
        calc_val = ws.cell(row=row_idx, column=overtime_calc_col_idx).value
        if calc_val is not None and calc_val != "" and calc_val != 0:
            ws.cell(row=row_idx, column=overtime_calc_col_idx).fill = orange_fill

        # 加班时长-上报时长 非空 → 绿色
        report_val = ws.cell(row=row_idx, column=overtime_report_col_idx).value
        if report_val is not None and report_val != "":
            ws.cell(row=row_idx, column=overtime_report_col_idx).fill = green_fill

    wb.save(TABLE1_FILE)

    print(f"生成完成：{TABLE1_FILE}")
    print(f"共 {len(result_df)} 人")
    print(result_df.to_string(index=False))


# ============================================================
# 表2
# ============================================================
def get_table2():
    """
    基于【工作安排表】生成【表2工作安排信息提取表】。

    从各项目行提取负责人及人员，结合请假/未到岗/休息名单，
    生成每人当天的工作安排状态和所属项目。
    """
    print("=" * 50)
    print("       生成【表2工作安排信息提取表】")
    print("=" * 50)

    # 1. 读取工作安排表
    if not WORK_PLAN_FILE.exists():
        raise FileNotFoundError(f"未找到工作安排表：{WORK_PLAN_FILE}")

    df = pd.read_excel(WORK_PLAN_FILE, header=HEADER_ROW2)

    # 规范化列名：去除空白字符（列名含多余空格，如"人     员"）
    df.columns = [re.sub(r"\s+", "", str(c)) for c in df.columns]

    print(f"读取工作安排表：共 {len(df)} 行数据")

    # 2. 收集人员信息
    #   person_info: { 姓名: {"出工安排": str, "项目": str} }
    person_info: dict[str, dict] = {}

    # --- 2a. 从常规项目行提取出工人员 ---
    for _, row in df.iterrows():
        project = row.get("项目", "")
        if pd.isna(project) or str(project).strip() == "":
            continue
        project = str(project).strip()

        # 跳过特殊状态行（后续单独处理）
        if project in STATUS_ROWS:
            continue

        leader = row.get("负责人", "")
        people_str = row.get("人员", "")

        # 收集该项目的所有人员
        all_people = []
        if not pd.isna(leader):
            all_people.extend(parse_people(str(leader)))
        if not pd.isna(people_str):
            all_people.extend(parse_people(str(people_str)))

        # 去重后登记
        for name in dict.fromkeys(all_people):
            if name in person_info:
                # 已在其他项目出现过，追加项目名
                existing_proj = person_info[name]["项目"]
                if project not in existing_proj.split("、"):  # 避免同一项目重复
                    person_info[name]["项目"] = f"{existing_proj}、{project}"
            else:
                person_info[name] = {"出工安排": "出工", "项目": project}

    # --- 2b. 从特殊状态行覆盖状态（人员名单在"车辆"列） ---
    for _, row in df.iterrows():
        project = row.get("项目", "")
        if pd.isna(project):
            continue
        project = str(project).strip()

        if project not in STATUS_ROWS:
            continue

        status = STATUS_ROWS[project]
        # 特殊状态行的人员名单写在"车辆"列
        people_str = row.get("车辆", "")
        all_people = parse_people(str(people_str) if not pd.isna(people_str) else "")

        for name in dict.fromkeys(all_people):
            person_info[name] = {"出工安排": status, "项目": ""}

    # 3. 生成表2
    rows = []
    for name in sorted(person_info.keys()):
        info = person_info[name]
        rows.append(
            {
                "姓名": name,
                "出工安排": info["出工安排"],
                "项目": info["项目"],
            }
        )

    columns = ["姓名", "出工安排", "项目"]
    result_df = pd.DataFrame(rows, columns=columns)
    result_df.to_excel(TABLE2_FILE, index=False)

    print(f"生成完成：{TABLE2_FILE}")
    print(f"共 {len(result_df)} 人")

    # 打印统计
    for status in ["出工", "请假", "外协未到岗", "安排休息"]:
        count = len(result_df[result_df["出工安排"] == status])
        if count > 0:
            print(f"  - {status}: {count} 人")

    print()
    print(result_df.to_string(index=False))


# ============================================================
# 表3
# ============================================================
def get_table3():
    """
    基于【output/00_程序人脸识别结果.xlsx】生成【表3出工照片人脸识别结果表】。

    从识别结果中解析每张照片的拍摄人姓名、拍摄时间、人脸序列号，
    并关联识别结果，输出为 Excel 表格文件。
    """
    print("=" * 50)
    print("       生成【表3出工照片人脸识别结果表】")
    print("=" * 50)

    # 1. 读取识别结果
    if not ATTENDANCE_FILE.exists():
        raise FileNotFoundError(f"未找到 00_程序人脸识别结果.xlsx：{ATTENDANCE_FILE}")

    df = pd.read_excel(ATTENDANCE_FILE)
    print(f"读取 00_程序人脸识别结果.xlsx：共 {len(df)} 条记录")

    # 2. 解析图片文件名
    #    xlsx 中"被识别图像名称"不含"(人脸N)"后缀，人脸序号单独在"人脸"列
    parsed = df["被识别图像名称"].apply(parse_filename)
    names, times, _ = zip(*parsed)

    # 3. 生成表3
    rows = []
    for name, t, seq, result in zip(
        names, times, df["人脸"].fillna(""), df["识别人名"].fillna("")
    ):
        rows.append(
            {
                "拍摄人姓名": name,
                "时间": t,
                "序列": seq,
                "识别结果": result,
            }
        )

    columns = ["拍摄人姓名", "时间", "序列", "识别结果"]
    result_df = pd.DataFrame(rows, columns=columns)
    result_df.to_excel(TABLE3_FILE, index=False)

    print(f"生成完成：{TABLE3_FILE}")
    print(f"共 {len(result_df)} 条记录")
    print(result_df.head(10).to_string(index=False))
    print("...")


# ============================================================
# 表4
# ============================================================
def get_table4():
    """
    基于【表3出工照片人脸识别结果表】生成【表4出工照片识别出工人表】。

    从表3中提取所有"识别出的人"以及拍摄人本人，
    按人汇总出现的时间点和拍摄人。
    """
    print("=" * 50)
    print("       生成【表4出工照片识别出工人表】")
    print("=" * 50)

    # 1. 读取表3
    if not TABLE3_FILE.exists():
        raise FileNotFoundError(f"未找到表3：{TABLE3_FILE}")

    df3 = pd.read_excel(TABLE3_FILE)
    print(f"读取表3：共 {len(df3)} 条记录")

    # 2. 汇总每个人的时间点和拍摄人
    #    { 姓名: { "times": set(), "photographers": set() } }
    person_records: dict[str, dict] = {}

    for _, row in df3.iterrows():
        photo_person = str(row["拍摄人姓名"]).strip()
        photo_time = str(row["时间"]).strip()
        result = str(row["识别结果"]).strip()

        # a. 拍摄人本人（只要有照片，拍摄人就在场）
        person_records.setdefault(photo_person, {"times": [], "photographers": []})
        if photo_time not in person_records[photo_person]["times"]:
            person_records[photo_person]["times"].append(photo_time)
        if photo_person not in person_records[photo_person]["photographers"]:
            person_records[photo_person]["photographers"].append(photo_person)

        # b. 识别结果中已识别的人
        if result not in ("Unknown", "未检测到人脸", NO_FACE_CONFIRMED, "", "nan"):
            person_records.setdefault(result, {"times": [], "photographers": []})
            if photo_time not in person_records[result]["times"]:
                person_records[result]["times"].append(photo_time)
            if photo_person not in person_records[result]["photographers"]:
                person_records[result]["photographers"].append(photo_person)

    # 3. 对每个人的时间排序
    for name in person_records:
        person_records[name]["times"].sort()
        person_records[name]["photographers"].sort()

    # 4. 固定5个时间列，预留扩展空间
    TIME_COLS = 5

    # 5. 生成表4
    rows = []
    for name in sorted(person_records.keys()):
        info = person_records[name]
        row_data = {"姓名": name}
        for i in range(TIME_COLS):
            col_name = f"时间{i + 1}"
            row_data[col_name] = info["times"][i] if i < len(info["times"]) else ""
        row_data["拍摄人"] = "、".join(info["photographers"])
        rows.append(row_data)

    columns = ["姓名"] + [f"时间{i + 1}" for i in range(TIME_COLS)] + ["拍摄人"]
    result_df = pd.DataFrame(rows, columns=columns)
    result_df.to_excel(TABLE4_FILE, index=False)

    actual_max = (
        max(len(v["times"]) for v in person_records.values()) if person_records else 0
    )
    print(f"生成完成：{TABLE4_FILE}")
    print(
        f"共 {len(result_df)} 人，实际最多 {actual_max} 个时间点（预留 {TIME_COLS} 列）"
    )
    print(result_df.to_string(index=False))


# ============================================================
# 表5（异常检验1：出工信息错报漏报）
# ============================================================
def get_table5():
    """
    对比 表1（台账申报）与 表4（人脸识别），生成 05_该日出工人员表。

      - 对比无问题 → 直接输出 output/05_该日出工人员表.xlsx（表1 内容）
      - 对比有问题 → 额外输出 output/核对信息错误文档1.xlsx 供人工核对
    """
    print("=" * 50)
    print("       生成【05_该日出工人员表】")
    print("=" * 50)

    # 1. 读取表1和表4
    if not TABLE1_FILE.exists():
        raise FileNotFoundError(f"未找到表1：{TABLE1_FILE}")
    if not TABLE4_FILE.exists():
        raise FileNotFoundError(f"未找到表4：{TABLE4_FILE}")

    df1 = pd.read_excel(TABLE1_FILE)
    df4 = pd.read_excel(TABLE4_FILE)

    print(f"表1 出工人数: {len(df1)}")
    print(f"表4 出工人数: {len(df4)}")

    count_match = len(df1) == len(df4)

    # 2. 构建对比数据
    #    表1: {姓名: {"开工时间": str, "收工时间": str, "拍摄人": str}}
    #    表4: {姓名: {"times": set, "拍摄人": str}}
    table1_info = {}
    for _, row in df1.iterrows():
        name = str(row["姓名"]).strip()
        # 01 的拍摄人已拆为 开工拍摄人/收工拍摄人，合并供对比
        ph_parts = [row.get("开工拍摄人"), row.get("收工拍摄人")]
        ph = "、".join(
            dict.fromkeys(
                str(p).strip() for p in ph_parts
                if pd.notna(p) and str(p).strip()
            )
        )
        table1_info[name] = {
            "开工时间": row.get("开工汇报拍摄时间", ""),
            "收工时间": row.get("收工汇报拍摄时间", ""),
            "拍摄人": ph,
        }

    table4_info = {}
    for _, row in df4.iterrows():
        name = str(row["姓名"]).strip()
        time_cols = [c for c in df4.columns if c.startswith("时间")]
        times = collect_times(row, time_cols)
        table4_info[name] = {
            "times": times,
            "拍摄人": row.get("拍摄人", ""),
        }

    # 3. 姓名相近匹配：表1 上报名 ↔ 表4 识别名（如 许士详↔许士祥、韩承俊↔韩承峻）
    def _name_sim(a, b):
        return difflib.SequenceMatcher(None, a, b).ratio()

    t1_names = set(table1_info.keys())
    t4_names = set(table4_info.keys())
    t1_only = sorted(n for n in t1_names if n not in t4_names)
    t4_only = sorted(n for n in t4_names if n not in t1_names)
    name_map = {}  # 表1名 -> 表4名（相近匹配，05 用表4识别名）
    used_t4 = set()
    for n1 in t1_only:
        best, best_sim = None, 0.0
        for n4 in t4_only:
            if n4 in used_t4 or len(n1) != len(n4):
                continue
            diff = sum(1 for a, b in zip(n1, n4) if a != b)
            if diff > 1:  # 至多差 1 字
                continue
            s = _name_sim(n1, n4)
            if s > best_sim:
                best, best_sim = n4, s
        if best and best_sim >= 0.6:
            name_map[n1] = best
            used_t4.add(best)
    if name_map:
        print(f"  ✅ 姓名相近匹配 {len(name_map)} 对："
              + "、".join(f"{k}→{v}" for k, v in name_map.items()))

    # 4. 逐人对比（用规范名：被映射的表1名并入对应表4名）
    inverse_map = {v: k for k, v in name_map.items()}
    all_names = sorted(
        {name_map.get(n, n) for n in t1_names} | t4_names
    )

    rows = []
    normal_count = 0
    abnormal_count = 0
    abnormal_names = []

    for name in all_names:
        t1_key = inverse_map.get(name, name)  # 合并后该人的表1名
        in_t1 = t1_key in table1_info
        in_t4 = name in table4_info

        t1_start = ""
        t1_end = ""
        t1_photographer = ""
        t4_times_list = []
        t4_photographer = ""

        if in_t1:
            info = table1_info[t1_key]
            t1_start = (
                str(info["开工时间"])
                if pd.notna(info["开工时间"]) and str(info["开工时间"]) != ""
                else ""
            )
            t1_end = (
                str(info["收工时间"])
                if pd.notna(info["收工时间"]) and str(info["收工时间"]) != ""
                else ""
            )
            t1_photographer = str(info["拍摄人"]) if pd.notna(info["拍摄人"]) else ""

        if in_t4:
            info = table4_info[name]
            t4_times_list = sorted(info["times"])
            t4_photographer = str(info["拍摄人"]) if pd.notna(info["拍摄人"]) else ""

        # 判断结论
        conclusion = ""
        abnormal = False

        if in_t1 and in_t4:
            # 同时出现：检查表1的时间是否都在表4中存在
            t1_times = {t for t in [t1_start, t1_end] if t}
            missing_in_t4 = t1_times - info["times"]

            if not missing_in_t4:
                conclusion = "正常出工"
                normal_count += 1
            else:
                if len(t1_times) > len(info["times"]):
                    conclusion = (
                        "出工异常人工复核（表1时间多于表4，可能人脸未匹配上/出工多报）"
                    )
                else:
                    conclusion = (
                        "出工异常人工复核（表1时间与表4时间不匹配，可能出工错报）"
                    )
                abnormal = True
        elif in_t1 and not in_t4:
            conclusion = (
                "出工异常人工复核（表4未出现，可能人脸未匹配上/出工多报/出工错报）"
            )
            abnormal = True
        elif not in_t1 and in_t4:
            conclusion = "出工异常人工复核（表1未出现，可能人脸识别错误/出工漏报）"
            abnormal = True

        if abnormal:
            abnormal_count += 1
            abnormal_names.append(name)

        rows.append(
            {
                "姓名": name,
                "表1开工汇报拍摄时间": t1_start,
                "表1收工汇报拍摄时间": t1_end,
                "表1拍摄人": t1_photographer,
                "表4时间点": "、".join(t4_times_list),
                "表4拍摄人": t4_photographer,
                "判断结论": conclusion,
            }
        )

    # 4. 输出 05_该日出工人员表
    #    （表1 内容去掉校验列；姓名用表4识别名纠正相近匹配；
    #      开工/收工时间、项目名、拍摄人 优先用表4人脸识别的实际数据）
    drop_cols = ["是否开工汇报异常", "是否收工汇报异常", "拍摄地点", "存在问题"]
    df5_out = df1.drop(columns=[c for c in drop_cols if c in df1.columns])
    if name_map:
        df5_out["姓名"] = df5_out["姓名"].map(
            lambda n: name_map.get(str(n).strip(), n)
        )

    # 4.1 照片索引：拍摄时间 -> {工程名称, 汇报类型, 拍摄人}（用于给表4时间配项目）
    def _ptime(t):
        try:
            return pd.to_datetime(t).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return str(t).strip()

    photo_index = {}
    try:
        ledger = pd.read_excel(PHOTO_LEDGER_FILE, header=HEADER_ROW)
        for _, lr in ledger.iterrows():
            t = lr.get("拍摄时间")
            if pd.isna(t):
                continue
            k = _ptime(t)
            if k not in photo_index:
                photo_index[k] = {
                    "工程名称": str(lr.get("工程名称", "")).strip() if pd.notna(lr.get("工程名称")) else "",
                    "汇报类型": str(lr.get("汇报类型", "")).strip() if pd.notna(lr.get("汇报类型")) else "",
                    "拍摄人": str(lr.get("拍摄人", "")).strip() if pd.notna(lr.get("拍摄人")) else "",
                }
    except Exception:
        photo_index = {}

    # 4.2 表4 每人实际开/收工：按识别时间 → 台账的 汇报类型/工程名称/拍摄人
    t4_person = {}
    for _, r in df4.iterrows():
        name = str(r["姓名"]).strip()
        times = [str(r[c]).strip() for c in df4.columns
                 if c.startswith("时间") and pd.notna(r[c]) and str(r[c]).strip()]
        photographers = [p.strip() for p in str(r.get("拍摄人", "")).split("、") if p.strip()]
        start, end = "", ""
        start_proj, end_proj = "", ""
        start_ph, end_ph = "", ""
        for t in times:
            rec = photo_index.get(_ptime(t))
            if not rec:
                continue
            ph = rec["拍摄人"] or (photographers[0] if photographers else "")
            if rec["汇报类型"] == "开工汇报":
                if not start or t < start:
                    start, start_proj, start_ph = t, rec["工程名称"], ph
            elif rec["汇报类型"] == "收工汇报":
                if not end or t > end:
                    end, end_proj, end_ph = t, rec["工程名称"], ph
        t4_person[name] = {
            "开工时间": start, "开工项目名": start_proj, "开工拍摄人": start_ph,
            "收工时间": end, "收工项目名": end_proj, "收工拍摄人": end_ph,
        }

    # 4.3 05 的开/收工字段完全基于表4（人脸识别）实际数据：
    #     被识别到的按表4填（表4未确认的置空），未识别到（仅台账误报）的全部置空
    clear_fields = ["开工汇报拍摄时间", "开工项目名", "开工拍摄人",
                    "收工汇报拍摄时间", "收工项目名", "收工拍摄人"]
    for idx, row in df5_out.iterrows():
        nm = str(row["姓名"]).strip()
        d = t4_person.get(nm)
        if not d:
            # 未被人脸识别（照片台账误报）→ 开/收工字段清空
            for f in clear_fields:
                df5_out.at[idx, f] = ""
        else:
            df5_out.at[idx, "开工汇报拍摄时间"] = d["开工时间"]
            df5_out.at[idx, "开工项目名"] = d["开工项目名"]
            df5_out.at[idx, "开工拍摄人"] = d["开工拍摄人"]
            df5_out.at[idx, "收工汇报拍摄时间"] = d["收工时间"]
            df5_out.at[idx, "收工项目名"] = d["收工项目名"]
            df5_out.at[idx, "收工拍摄人"] = d["收工拍摄人"]
        # 全天状态 / 加班时长 基于最新时间重算
        kg = df5_out.at[idx, "开工项目名"]
        sg = df5_out.at[idx, "收工项目名"]
        df5_out.at[idx, "全天状态"] = (
            "全天一致" if (kg and sg and kg == sg)
            else "半天换场" if (kg and sg and kg != sg)
            else ""
        )
        ot = df5_out.at[idx, "收工汇报拍摄时间"]
        if ot:
            df5_out.at[idx, "加班时长-基于收工时间测算"] = calc_overtime(ot)

    df5_out.to_excel(TABLE5_FILE, index=False)
    print(f"✅ 已生成 05_该日出工人员表: {TABLE5_FILE}")
    if t4_person:
        print(f"  （表4实际出工数据覆盖 {len(t4_person)} 人的开/收工字段）")

    # 5. 汇总报告
    total = len(all_names)
    print(f"\n{'='*50}")
    print(f" 表1人数: {len(df1)}  |  表4人数: {len(df4)}")
    print(f" 正常出工: {normal_count} 人  |  异常: {abnormal_count} 人")
    print(f" 总对比人数: {total} 人")
    print(f" 总数一致: {'是' if count_match else '否'}")
    if abnormal_names:
        print(f" 异常人员: {', '.join(abnormal_names)}")

    # 6. 有异常 → 生成核对信息错误文档1
    abnormal_rows = [r for r in rows if "异常" in r["判断结论"]]
    if abnormal_rows:
        _generate_error_doc1(
            rows, abnormal_rows, df1, df4, normal_count, abnormal_count, count_match
        )
        print(f"⚠️ 有 {abnormal_count} 处异常，请核对: {TABLE5_ERROR_FILE}")
    else:
        print("✅ 检验正常 - 无需人工复核")

    # 7. 返回汇总信息（供 web 层判断是否弹窗确认，main.py 忽略返回值）
    return {
        "normal": normal_count,
        "abnormal": abnormal_count,
        "error_doc": TABLE5_ERROR_FILE if abnormal_rows else None,
        "table_file": TABLE5_FILE,
    }


def _generate_error_doc1(
    rows, abnormal_rows, df1, df4, normal_count, abnormal_count, count_match
):
    """生成 核对信息错误文档1.xlsx（异常明细 + 复核建议）。"""
    wb_review = Workbook()
    ws = wb_review.active
    ws.title = "核对信息错误清单"

    # 样式定义
    header_font = Font(name="微软雅黑", bold=True, size=11)
    header_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    title_font = Font(name="微软雅黑", bold=True, size=14)
    text_font = Font(name="微软雅黑", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 汇总信息
    ws.merge_cells("A1:J1")
    ws["A1"] = "【核对信息错误文档1】出工信息错报漏报"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    ws.merge_cells("A2:J2")
    ws["A2"] = (
        f"表1人数：{len(df1)}  |  表4人数：{len(df4)}  |  正常：{normal_count}人  |  异常：{abnormal_count}人  |  总数一致：{'是' if count_match else '否'}"
    )
    ws["A2"].font = Font(name="微软雅黑", size=10)
    ws["A2"].alignment = left_align

    # 统计各类异常
    cat_t4_not_found = [r for r in abnormal_rows if "表4未出现" in r["判断结论"]]
    cat_t1_not_found = [r for r in abnormal_rows if "表1未出现" in r["判断结论"]]
    cat_time_mismatch = [
        r
        for r in abnormal_rows
        if "时间" in r["判断结论"] and "未出现" not in r["判断结论"]
    ]

    row_idx = 4
    for cat_list, desc in [
        (
            cat_t4_not_found,
            "表4未出现（表1有、表4无，可能人脸未匹配上/出工多报/出工错报）",
        ),
        (cat_t1_not_found, "表1未出现（表4有、表1无，可能人脸识别错误/出工漏报）"),
        (cat_time_mismatch, "时间不匹配（表1与表4时间不一致）"),
    ]:
        if cat_list:
            ws.merge_cells(f"A{row_idx}:J{row_idx}")
            ws[f"A{row_idx}"] = f"▶ {desc}：{len(cat_list)}人"
            ws[f"A{row_idx}"].font = Font(
                name="微软雅黑", bold=True, size=10, color="C00000"
            )
            row_idx += 1

    # 明细表头
    row_idx += 1
    detail_headers = [
        "序号",
        "姓名",
        "异常类别",
        "表1开工汇报拍摄时间",
        "表1收工汇报拍摄时间",
        "表1拍摄人",
        "表4时间点",
        "表4拍摄人",
    ]
    for col_idx, h in enumerate(detail_headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    # 明细数据
    for i, r in enumerate(abnormal_rows, 1):
        row_idx += 1
        if "表4未出现" in r["判断结论"]:
            category = "表4未出现"
        elif "表1未出现" in r["判断结论"]:
            category = "表1未出现"
        elif (
            len(r.get("表1开工汇报拍摄时间", "")) > 0
            or len(r.get("表1收工汇报拍摄时间", "")) > 0
        ):
            category = "时间不匹配"
        else:
            category = "其他异常"

        vals = [
            i,
            r["姓名"],
            category,
            r.get("表1开工汇报拍摄时间", ""),
            r.get("表1收工汇报拍摄时间", ""),
            r.get("表1拍摄人", ""),
            r.get("表4时间点", ""),
            r.get("表4拍摄人", ""),
        ]
        for col_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.font = text_font
            cell.border = thin_border
            cell.alignment = left_align if col_idx >= 4 else center_align

    # 列宽
    col_widths = [6, 10, 12, 22, 22, 12, 30, 12]
    for col_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    wb_review.save(TABLE5_ERROR_FILE)
    print(f" 核对信息错误文档1: {TABLE5_ERROR_FILE}")


# ============================================================
# 表6（对比 05_ 与表2，生成 06_全体人员出工情况表）
# ============================================================
def _project_matches(t2_proj: str, t5_kaigong: str, t5_shougong: str) -> bool:
    """判断表2安排项目是否与05_开工/收工项目匹配（模糊匹配）。"""
    if not t2_proj:
        return True

    t2_clean = _normalize_project(t2_proj)

    targets = []
    if t5_kaigong:
        targets.append(_normalize_project(t5_kaigong))
    if t5_shougong:
        targets.append(_normalize_project(t5_shougong))

    if not targets:
        return True

    for t in targets:
        if t2_clean in t or t in t2_clean:
            return True
        t2_key = _extract_key(t2_clean)
        t5_key = _extract_key(t)
        if t2_key and t5_key and (t2_key in t5_key or t5_key in t2_key):
            return True

    return False


def _normalize_project(s: str) -> str:
    """清洗项目名：统一小写、去除所有空格（含内部空格）。"""
    s = s.strip().lower()
    s = re.sub(r"\s+", "", s)
    return s


def _extract_key(s: str) -> str:
    """提取项目名关键词（变电站名主体），去除电压等级、空格与常见噪声。

    覆盖人工录入常见错误：电压等级大小写 / 缺单位（220kv、220V、220千伏）、
    多打空格、多写"项目 / 工程 / 变电站"等后缀 —— 只要核心站点名一致即视为匹配。
    """
    # 去电压等级前缀（220kv / 110KV / 220V / 500kV / 220千伏 等）
    s = re.sub(r"\d+\s*(?:k|m|g)?v", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\d+\s*(?:千伏|万伏|伏)", "", s)
    # 去空格
    s = re.sub(r"\s+", "", s)
    # 去常见后缀噪声
    for noise in ("变电站", "变电所", "项目部", "项目", "工程"):
        s = s.replace(noise, "")
    # 去"变"，保留站名主体（如 郭巷变 → 郭巷）
    s = s.replace("变", "")
    return s if len(s) >= 2 else ""


def _find_canonical_project(raw, canonical_projects):
    """在表2规范项目列表中，找到与 05_ 项目名实质匹配的规范写法。

    - 找到匹配 → 返回规范写法（如 220kv 郭巷变 → 220kV郭巷变）
    - 找不到匹配（如"公司"）→ 保留原内容，仅去除空格
    """
    raw_s = str(raw).strip()
    if not raw_s:
        return ""
    best, best_len = None, -1
    for c in canonical_projects:
        if _project_matches(c, raw_s, ""):
            if c == raw_s:
                return c
            klen = len(_extract_key(c))
            if klen > best_len:
                best, best_len = c, klen
    if best:
        return best
    # 未匹配：保留原内容，仅去除空格
    return re.sub(r"\s+", "", raw_s)


def get_table6():
    """
    对比 05_该日出工人员表（实际）与 表2（计划），生成 06_全体人员出工情况表。

      - 对比无问题 → 直接输出 output/06_全体人员出工情况表.xlsx
      - 对比有问题 → 额外输出 output/核对信息错误文档2.xlsx 供人工核对
      - 06_ 的开工/收工项目简称 通过 002项目信息表（000_项目信息表）匹配
    """
    print("=" * 50)
    print("       生成【06_全体人员出工情况表】")
    print("=" * 50)

    # 1. 读取 05_该日出工人员表、表2
    if not TABLE5_FILE.exists():
        raise FileNotFoundError(f"未找到 05_该日出工人员表：{TABLE5_FILE}")
    if not TABLE2_FILE.exists():
        raise FileNotFoundError(f"未找到表2：{TABLE2_FILE}")

    df5 = pd.read_excel(TABLE5_FILE)
    df2 = pd.read_excel(TABLE2_FILE)

    print(f"05_该日出工人员数: {len(df5)}")
    print(f"表2 安排人数: {len(df2)}")

    # 2. 构建对比数据
    #    表2: {姓名: {"出工安排": str, "项目": str}}
    table2_info = {}
    for _, row in df2.iterrows():
        name = str(row["姓名"]).strip()
        table2_info[name] = {
            "出工安排": str(row.get("出工安排", "")).strip(),
            "项目": str(row.get("项目", "")) if pd.notna(row.get("项目")) else "",
        }

    #    05_: {姓名: {"开工项目名": str, "收工项目名": str, "全天状态": str}}
    table5_info = {}
    for _, row in df5.iterrows():
        name = str(row["姓名"]).strip()
        table5_info[name] = {
            "开工项目名": (
                str(row.get("开工项目名", ""))
                if pd.notna(row.get("开工项目名"))
                else ""
            ),
            "收工项目名": (
                str(row.get("收工项目名", ""))
                if pd.notna(row.get("收工项目名"))
                else ""
            ),
            "全天状态": (
                str(row.get("全天状态", "")) if pd.notna(row.get("全天状态")) else ""
            ),
        }

    # 3. 逐人对比（计划 vs 实际）
    all_names = sorted(set(table2_info.keys()) | set(table5_info.keys()))

    rows = []
    normal_count = 0
    abnormal_count = 0
    abnormal_details = []

    for name in all_names:
        in_t2 = name in table2_info
        in_t5 = name in table5_info

        t2_arrange = ""
        t2_project = ""
        t5_kaigong_proj = ""
        t5_shougong_proj = ""
        t5_status = ""

        if in_t2:
            info = table2_info[name]
            t2_arrange = info["出工安排"]
            t2_project = info["项目"]

        if in_t5:
            info = table5_info[name]
            t5_kaigong_proj = info["开工项目名"]
            t5_shougong_proj = info["收工项目名"]
            t5_status = info["全天状态"]

        # 判断结论
        conclusion = ""
        abnormal = False
        # 该人是否实际被人脸识别到岗（05 开/收工项目名非空才算真正出工）
        actually_worked = bool(t5_kaigong_proj or t5_shougong_proj)

        if in_t2 and in_t5:
            if t2_arrange == "出工":
                proj_match = _project_matches(
                    t2_project, t5_kaigong_proj, t5_shougong_proj
                )
                if proj_match:
                    conclusion = "正常出工"
                    normal_count += 1
                else:
                    conclusion = "出工异常人工复核（安排项目与实际项目不匹配）"
                    abnormal = True
            elif t2_arrange in ("安排休息", "请假", "外协未到岗"):
                if actually_worked:
                    conclusion = (
                        "出工异常人工复核（安排为「" + t2_arrange + "」但实际有出工记录）"
                    )
                    abnormal = True
                else:
                    conclusion = "正常（安排「" + t2_arrange + "」，实际未识别到出工，一致）"
                    normal_count += 1
            else:
                conclusion = "出工异常人工复核（安排状态未知，需核实）"
                abnormal = True
        elif in_t2 and not in_t5:
            if t2_arrange == "出工":
                conclusion = (
                    "出工异常人工复核（安排了出工但无实际出工记录，可能漏报/缺勤）"
                )
                abnormal = True
            else:
                conclusion = "正常（安排「" + t2_arrange + "」，无出工记录，一致）"
                normal_count += 1
        elif not in_t2 and in_t5:
            if actually_worked:
                conclusion = (
                    "出工异常人工复核（表5有出工记录但表2无此人的安排信息，可能安排遗漏）"
                )
                abnormal = True
            else:
                conclusion = "正常（表5在列但未识别到实际出工，可能为台账误报）"
                normal_count += 1

        if abnormal:
            abnormal_count += 1

        row_data = {
            "姓名": name,
            "表2出工安排": t2_arrange,
            "表2安排项目": t2_project,
            "05_开工项目名": t5_kaigong_proj,
            "05_收工项目名": t5_shougong_proj,
            "05_全天状态": t5_status,
            "判断结论": conclusion,
        }
        rows.append(row_data)
        if abnormal:
            abnormal_details.append(row_data)

    # 4. 生成 06_全体人员出工情况表（在 05_ 基础上补充项目简称）
    exact_lookup, all_entries, name_entries = build_project_mapping()
    print(f"加载002项目信息表: {len(exact_lookup)} 条记录")

    # 删除加班时长测算列
    drop_cols = ["加班时长-基于收工时间测算", "加班时长-上报时长"]
    df6 = df5.drop(columns=[c for c in drop_cols if c in df5.columns])

    # 修正项目名为表2规范写法：先收集表2安排项目的标准写法
    canonical_projects = []
    _seen = set()
    for _info in table2_info.values():
        _p = str(_info.get("项目", "") or "").strip()
        if _p and _p not in _seen:
            _seen.add(_p)
            canonical_projects.append(_p)

    # 05_ 的开工/收工项目名 → 规范写法（实质匹配则替换为表2标准写法；
    # 未匹配（如"公司"）保留原内容，仅去除空格）
    for col in ["开工项目名", "收工项目名"]:
        if col in df6.columns:
            df6[col] = df6[col].astype(str).map(
                lambda v: _find_canonical_project(v, canonical_projects)
            )

    # 匹配项目简称
    kaigong_shorts = []
    kaigong_remarks = []
    shougong_shorts = []
    shougong_remarks = []
    for _, row in df6.iterrows():
        kg_short, _, kg_remark = match_project(
            row.get("开工项目名", ""), "开工项目名", exact_lookup, all_entries, name_entries
        )
        kaigong_shorts.append(kg_short)
        kaigong_remarks.append(kg_remark)

        sg_short, _, sg_remark = match_project(
            row.get("收工项目名", ""), "收工项目名", exact_lookup, all_entries, name_entries
        )
        shougong_shorts.append(sg_short)
        shougong_remarks.append(sg_remark)

    # 组装列顺序：在开工/收工项目名后插入简称，最后加备注
    cols = df6.columns.tolist()
    kg_idx = cols.index("开工项目名")
    sg_idx = cols.index("收工项目名")

    new_data = {c: df6[c].values for c in df6.columns}
    new_data["开工项目简称"] = kaigong_shorts
    new_data["收工项目简称"] = shougong_shorts
    new_data["备注"] = ""

    new_cols = []
    for i, c in enumerate(cols):
        new_cols.append(c)
        if i == kg_idx:
            new_cols.append("开工项目简称")
        if i == sg_idx:
            new_cols.append("收工项目简称")
    new_cols.append("备注")

    df6 = pd.DataFrame(new_data, columns=new_cols)

    # 合并备注
    combined_remarks = []
    for kg_r, sg_r in zip(kaigong_remarks, shougong_remarks):
        parts = []
        if kg_r:
            parts.append(kg_r)
        if sg_r:
            parts.append(sg_r)
        combined_remarks.append("；".join(parts))
    df6["备注"] = combined_remarks

    # 追加非出工人员
    non_work_status = ["请假", "外协未到岗", "安排休息"]
    non_work_df = df2[df2["出工安排"].isin(non_work_status)].copy()
    STATUS_TO_SHORT = {"安排休息": "休", "外协未到岗": "空", "请假": "假"}

    append_rows = []
    for _, row in non_work_df.iterrows():
        name = str(row["姓名"]).strip()
        status = str(row["出工安排"]).strip()
        append_rows.append(
            {
                "姓名": name,
                "开工项目简称": STATUS_TO_SHORT.get(status, ""),
                "收工项目简称": STATUS_TO_SHORT.get(status, ""),
                "全天状态": status,
            }
        )

    if append_rows:
        append_df = pd.DataFrame(append_rows)
        for c in df6.columns:
            if c not in append_df.columns:
                append_df[c] = ""
        append_df = append_df[df6.columns]
        df6 = pd.concat([df6, append_df], ignore_index=True)
        print(f"追加非出工人员: {len(append_rows)} 人")

    # 输出 06_ + 高亮
    df6.to_excel(TABLE6_FILE, index=False)
    wb = load_workbook(TABLE6_FILE)
    ws = wb.active
    highlight_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )
    kg_short_col = new_cols.index("开工项目简称") + 1
    sg_short_col = new_cols.index("收工项目简称") + 1
    remark_col = new_cols.index("备注") + 1

    for row_idx in range(2, ws.max_row + 1):
        kg_val = ws.cell(row=row_idx, column=kg_short_col).value
        if not kg_val or str(kg_val).strip() == "":
            ws.cell(row=row_idx, column=kg_short_col).fill = highlight_fill
        sg_val = ws.cell(row=row_idx, column=sg_short_col).value
        if not sg_val or str(sg_val).strip() == "":
            ws.cell(row=row_idx, column=sg_short_col).fill = highlight_fill
        remark_val = ws.cell(row=row_idx, column=remark_col).value
        if remark_val and str(remark_val).strip():
            ws.cell(row=row_idx, column=remark_col).fill = highlight_fill

    wb.save(TABLE6_FILE)
    print(f"✅ 已生成 06_全体人员出工情况表: {TABLE6_FILE}")

    # 5. 汇总报告
    print(f"\n{'='*50}")
    print(f" 表2安排人数: {len(df2)}  |  05_实际出工人数: {len(df5)}")
    print(f" 正常: {normal_count} 人  |  异常: {abnormal_count} 人")
    print(f" 总对比人数: {len(all_names)} 人")

    # 6. 有异常 → 生成核对信息错误文档2
    if abnormal_details:
        _generate_error_doc2(df2, df5, normal_count, abnormal_count, abnormal_details)
        print(f"⚠️ 有 {abnormal_count} 处异常，请核对: {TABLE6_ERROR_FILE}")
    else:
        print("✅ 检验正常 - 无需人工复核")

    # 7. 返回汇总信息（供 web 层判断是否弹窗确认，main.py 忽略返回值）
    return {
        "normal": normal_count,
        "abnormal": abnormal_count,
        "error_doc": TABLE6_ERROR_FILE if abnormal_details else None,
        "table_file": TABLE6_FILE,
    }


def _generate_error_doc2(df2, df5, normal_count, abnormal_count, abnormal_details):
    """生成 核对信息错误文档2.xlsx（异常明细 + 复核建议）。"""
    wb_review = Workbook()
    ws = wb_review.active
    ws.title = "核对信息错误清单"

    header_font = Font(name="微软雅黑", bold=True, size=11)
    header_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    title_font = Font(name="微软雅黑", bold=True, size=14)
    text_font = Font(name="微软雅黑", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A1:H1")
    ws["A1"] = "【核对信息错误文档2】出工状态核对"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"表2安排人数：{len(df2)}  |  05_实际出工人数：{len(df5)}  |  "
        f"正常：{normal_count}人  |  异常：{abnormal_count}人"
    )
    ws["A2"].font = Font(name="微软雅黑", size=10)
    ws["A2"].alignment = left_align

    cat_scheduled_absent = [
        r for r in abnormal_details if "安排了出工但无实际出工记录" in r["判断结论"]
    ]
    cat_rested_present = [
        r
        for r in abnormal_details
        if "安排为" in r["判断结论"] and "实际有出工" in r["判断结论"]
    ]
    cat_no_schedule = [
        r for r in abnormal_details if "无此人的安排信息" in r["判断结论"]
    ]
    cat_project_mismatch = [
        r for r in abnormal_details if "项目不匹配" in r["判断结论"]
    ]
    cat_other = [
        r
        for r in abnormal_details
        if r not in cat_scheduled_absent
        and r not in cat_rested_present
        and r not in cat_no_schedule
        and r not in cat_project_mismatch
    ]

    row_idx = 4
    categories = [
        (cat_scheduled_absent, "安排出工但无实际出工记录（可能漏报/缺勤）"),
        (
            cat_rested_present,
            "安排休息/请假/外协未到岗但实际有出工记录（安排与实际不符）",
        ),
        (cat_no_schedule, "05_有出工记录但表2无安排信息（可能安排遗漏）"),
        (cat_project_mismatch, "出工但安排项目与实际项目不匹配（可能项目错报）"),
        (cat_other, "其他异常"),
    ]
    for cat_list, desc in categories:
        if cat_list:
            ws.merge_cells(f"A{row_idx}:H{row_idx}")
            ws[f"A{row_idx}"] = f"▶ {desc}：{len(cat_list)}人"
            ws[f"A{row_idx}"].font = Font(
                name="微软雅黑", bold=True, size=10, color="C00000"
            )
            row_idx += 1

    row_idx += 1
    detail_headers = [
        "序号",
        "姓名",
        "异常类别",
        "表2出工安排",
        "表2安排项目",
        "05_开工项目名",
        "05_收工项目名",
    ]
    for col_idx, h in enumerate(detail_headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    for i, r in enumerate(abnormal_details, 1):
        row_idx += 1
        judgement = r["判断结论"]

        if "安排了出工但无实际出工记录" in judgement:
            category = "安排出工未出工"
        elif "安排为" in judgement and "实际有出工" in judgement:
            category = "安排休息但有出工"
        elif "无此人的安排信息" in judgement:
            category = "出工但无安排"
        elif "项目不匹配" in judgement:
            category = "项目不匹配"
        else:
            category = "其他异常"

        vals = [
            i,
            r["姓名"],
            category,
            r.get("表2出工安排", ""),
            r.get("表2安排项目", ""),
            r.get("05_开工项目名", ""),
            r.get("05_收工项目名", ""),
        ]
        for col_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.font = text_font
            cell.border = thin_border
            cell.alignment = left_align if col_idx >= 4 else center_align

    col_widths = [6, 10, 16, 14, 16, 16, 16]
    for col_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    wb_review.save(TABLE6_ERROR_FILE)
    print(f" 核对信息错误文档2: {TABLE6_ERROR_FILE}")


# ============================================================
# 项目简称匹配（002项目信息表，供 get_table6 使用）
# ============================================================
def normalize(s: str) -> str:
    """清洗项目名：去空格、统一kv→kV、小写、去末尾"项目"后缀。"""
    s = re.sub(r"\s+", "", str(s).strip())
    s = s.lower()
    s = s.replace("kv", "kV")
    if s.endswith("项目"):
        s = s[:-2]
    return s


def extract_key(s: str) -> str:
    """提取项目名关键词（去掉电压等级、'变'字等）。"""
    s = re.sub(r"\d+kv\s*", "", s, flags=re.IGNORECASE)
    s = s.replace("变", "").replace("项目", "").replace("部", "")
    s = re.sub(r"\s+", "", s)
    return s.strip() if len(s.strip()) >= 2 else ""


def build_project_mapping():
    """从002项目信息表构建映射。

    Returns:
        exact_lookup: {normalized工作安排简称: (原始工作安排简称, 出勤统计简称)}
        all_entries: [(原始工作安排简称, 出勤统计简称, 工作安排简称_norm, 提取关键词)]
        name_entries: [(原始项目名称, 出勤统计简称, normalized项目名称)] —— 供按"项目名称"回退匹配
                     （如"…石湖变等13座…"这类长项目名在工作安排简称里查不到，但其项目名称列有完整名）
    """
    df_002 = pd.read_excel(PROJECT_INFO_FILE)

    exact_lookup = {}
    all_entries = []
    name_entries = []

    for _, row in df_002.iterrows():
        key = row.get("工作安排简称", "")
        val = row.get("出勤统计简称", "")

        if pd.notna(key) and str(key).strip():
            k = str(key).strip()
            k_norm = normalize(k)
            v = str(val).strip() if pd.notna(val) else ""
            exact_lookup[k_norm] = (k, v)
            all_entries.append((k, v, k_norm, extract_key(k_norm)))

        pname = row.get("项目名称", "")
        if pd.notna(pname) and str(pname).strip():
            p = str(pname).strip()
            p_norm = normalize(p)
            v = str(val).strip() if pd.notna(val) else ""
            name_entries.append((p, v, p_norm))

    return exact_lookup, all_entries, name_entries


def match_project(proj_raw, column_label, exact_lookup, all_entries, name_entries=None):
    """将项目名匹配002项目信息表（工作安排简称精确匹配 → 项目名称回退匹配 → 最相似记录备注）。

    Args:
        proj_raw: 05_中的项目名（开工或收工）
        column_label: "开工项目名" 或 "收工项目名"，用于备注
        exact_lookup: 工作安排简称精确匹配字典
        all_entries: 所有002条目（工作安排简称）
        name_entries: [(项目名称, 出勤统计简称, 项目名称_norm)]，供按项目名称回退匹配

    Returns:
        (简称str, 是否高亮bool, 备注str)
    """
    if not proj_raw or pd.isna(proj_raw) or str(proj_raw).strip() == "":
        return "", True, f"{column_label}为空"

    proj_clean = str(proj_raw).strip()
    proj_norm = normalize(proj_clean)

    # --- 精确匹配（工作安排简称） ---
    if proj_norm in exact_lookup:
        original_key, short_stat = exact_lookup[proj_norm]
        if short_stat:
            return short_stat, False, ""
        else:
            return (
                "",
                True,
                f"{column_label}「{proj_clean}」在002项目信息表中出勤统计简称为空，需人工补充",
            )

    # --- 按项目名称回退匹配（等变电站等长项目名在"工作安排简称"列查不到，但项目名称列有完整名） ---
    for pname, short_stat, p_norm in (name_entries or []):
        if not p_norm:
            continue
        if p_norm == proj_norm or p_norm in proj_norm or proj_norm in p_norm:
            if short_stat:
                return short_stat, False, ""
            break

    # --- 无法精确匹配，尝试查找最相似记录生成备注 ---
    best_match = _find_closest(proj_norm, proj_clean, all_entries)
    if best_match:
        orig_key, short_stat, _, _ = best_match
        remark = (
            f"{column_label}「{proj_clean}」在002项目信息表中无法精确匹配，"
            f"最接近记录为「{orig_key}」"
        )
        if short_stat:
            remark += f"（出勤统计简称：「{short_stat}」）"
        remark += "，请人工确认"
        return "", True, remark
    else:
        return (
            "",
            True,
            f"{column_label}「{proj_clean}」在002项目信息表中未找到对应记录，请人工确认",
        )


def _find_closest(proj_norm, proj_clean, all_entries):
    """在002条目中查找与输入项目名最相似的记录。"""
    proj_key = extract_key(proj_norm)
    if not proj_key:
        return None

    # 关键词包含匹配
    for entry in all_entries:
        _, _, _, entry_key = entry
        if entry_key and proj_key and (proj_key in entry_key or entry_key in proj_key):
            return entry

    # 部分字符串包含
    for entry in all_entries:
        entry_norm = entry[2]
        if len(proj_norm) >= 3 and len(entry_norm) >= 3:
            if proj_norm[:3] == entry_norm[:3]:
                return entry
            if proj_norm in entry_norm or entry_norm in proj_norm:
                return entry

    return None


# ============================================================
# 表7
# ============================================================
def get_table7():
    """
    基于【06_全体人员出工情况表】和【03_人员分类表】生成【07_表7出工地点及时长统计表】。

    数据源：06_全体人员出工情况表（get_table6 生成）
      - 类别 / 所属人员：对应 03_人员分类表
      - 上午/下午出工地点：06_的开工/收工项目简称
      - 加班时长：06_的加班时长确认-人工审核
    """
    print("=" * 50)
    print("  生成【07_表7出工地点及时长统计表】")
    print("=" * 50)

    # 1. 读取数据源
    df6 = pd.read_excel(TABLE6_FILE)
    df_001 = pd.read_excel(CLASSIFY_FILE)

    print(f"表6人数: {len(df6)}")
    print(f"001人员分类表人数: {len(df_001)}")

    # 2. 构建001查找字典
    person_info = {}
    for _, row in df_001.iterrows():
        name = str(row["姓名"]).strip()
        person_info[name] = {
            "类别": str(row.get("类别", "")) if pd.notna(row.get("类别")) else "",
            "所属人员": (
                str(row.get("所属人员", "")) if pd.notna(row.get("所属人员")) else ""
            ),
        }

    # 3. 逐人构建表7数据
    rows = []
    matched = 0
    unmatched = 0

    for _, row in df6.iterrows():
        name = str(row["姓名"]).strip()

        # 类别和所属人员
        info = person_info.get(name, {})
        # 精确匹配失败时，尝试字形变体映射
        if not info:
            variant_name = TABLE7_NAME_VARIANT_MAP.get(name, "")
            if variant_name:
                info = person_info.get(variant_name, {})
        category = info.get("类别", "")
        belong_to = info.get("所属人员", "")

        # 上午/下午出工地点
        morning_loc = row.get("开工项目简称", "")
        if pd.isna(morning_loc):
            morning_loc = ""
        afternoon_loc = row.get("收工项目简称", "")
        if pd.isna(afternoon_loc):
            afternoon_loc = ""

        # 加班时长
        overtime = row.get("加班时长确认-人工审核", "")
        if pd.isna(overtime):
            overtime = ""

        # 异常提醒汇总
        alerts = []
        if not category:
            unmatched += 1
            alerts.append("未在001人员分类表中找到")
            print(f"  注意: {name} 未在001人员分类表中找到")
        else:
            matched += 1

        if not morning_loc and not afternoon_loc:
            alerts.append("出工地点为空")
        elif not morning_loc:
            alerts.append("上午出工地点为空")
        elif not afternoon_loc:
            alerts.append("下午出工地点为空")

        alert_msg = "；".join(alerts)

        rows.append(
            {
                "姓名": name,
                "类别": category,
                "所属人员": belong_to,
                "上午出工地点": morning_loc,
                "下午出工地点": afternoon_loc,
                "加班时长": overtime,
                "异常提醒": alert_msg,
            }
        )

    # 4. 输出
    columns = [
        "姓名",
        "类别",
        "所属人员",
        "上午出工地点",
        "下午出工地点",
        "加班时长",
        "异常提醒",
    ]
    df7 = pd.DataFrame(rows, columns=columns)
    df7.to_excel(TABLE7_FILE, index=False)

    print(f"\n匹配成功: {matched} 人  |  未匹配: {unmatched} 人")
    print(f"总行数: {len(df7)}")
    print(f"输出文件: {TABLE7_FILE}")


# ============================================================
# 表8 / 表9 / 表10
# ============================================================
def read_table7():
    """读取表7，返回 {姓名: {类别, 所属人员, 上午出工地点, 下午出工地点, 加班时长, 异常提醒}}"""
    df7 = pd.read_excel(TABLE7_FILE)
    data = {}
    for _, row in df7.iterrows():
        name = str(row["姓名"]).strip()
        cat = str(row["类别"]).strip() if pd.notna(row.get("类别")) else ""
        bel = str(row["所属人员"]).strip() if pd.notna(row.get("所属人员")) else ""
        am = (
            str(row["上午出工地点"]).strip()
            if pd.notna(row.get("上午出工地点"))
            else ""
        )
        pm = (
            str(row["下午出工地点"]).strip()
            if pd.notna(row.get("下午出工地点"))
            else ""
        )
        ot = row.get("加班时长", "")
        if pd.isna(ot):
            ot = ""
        else:
            ot_str = str(ot).strip()
            # 整数则去掉小数
            try:
                if float(ot) == int(float(ot)):
                    ot = str(int(float(ot)))
                else:
                    ot = ot_str
            except ValueError:
                ot = ot_str
        alert = (
            str(row.get("异常提醒", "")).strip()
            if pd.notna(row.get("异常提醒"))
            else ""
        )
        data[name] = {
            "类别": cat,
            "所属人员": bel,
            "上午出工地点": am,
            "下午出工地点": pm,
            "加班时长": ot,
            "异常提醒": alert,
        }
    return data


def lookup_table7(name, table7_data):
    """在表7数据中查找某人，支持名字变体匹配。
    返回 (data_dict, variant_note)
    data_dict 为 None 表示未找到
    """
    if name in table7_data:
        return table7_data[name], ""
    # 名字变体匹配
    for variant, standard in TABLE8_NAME_VARIANT_MAP.items():
        if name == standard and variant in table7_data:
            return table7_data[variant], "（姓名变体匹配）"
        if name == variant and standard in table7_data:
            return table7_data[standard], "（姓名变体匹配）"
    return None, ""


def get_person_rows(ws, name_col_idx, item_col_idx):
    """
    从已合并单元格的sheet中解析人员→行映射。
    返回 [(姓名, 起始行, {项目名: row_idx})]

    name_col_idx: 姓名列的1-based索引
    item_col_idx: 考勤项目/地点列的1-based索引
    """
    persons = []
    current_name = None
    current_start = None
    current_items = {}

    max_row = ws.max_row

    for r in range(2, max_row + 1):  # 跳过表头
        name_val = ws.cell(row=r, column=name_col_idx).value
        item_val = ws.cell(row=r, column=item_col_idx).value

        if name_val and str(name_val).strip():
            # 新人物开始
            if current_name is not None:
                persons.append((current_name, current_start, current_items))
            current_name = str(name_val).strip()
            current_start = r
            current_items = {}

        if item_val and current_name:
            item_str = str(item_val).strip()
            current_items[item_str] = r

    # 最后一个人
    if current_name is not None:
        persons.append((current_name, current_start, current_items))

    return persons


def fill_table8(wb, table7_data):
    """填表8工程考勤表（公司人员）"""
    ws = wb["表8工程考勤表模板"]

    # 列: A=序号, B=姓 名, C=考勤项目, D=出工情况, E=异常提醒(new)
    name_col = 2  # B
    item_col = 3  # C
    fill_col = 4  # D
    alert_col = 5  # E (new)

    # 添加异常提醒表头
    ws.cell(row=1, column=alert_col, value="异常提醒")
    # 复制表头样式
    header_style = ws.cell(row=1, column=fill_col)
    alert_header = ws.cell(row=1, column=alert_col)
    if header_style.font:
        alert_header.font = copy(header_style.font)
    if header_style.fill:
        alert_header.fill = copy(header_style.fill)
    if header_style.alignment:
        alert_header.alignment = copy(header_style.alignment)
    if header_style.border:
        alert_header.border = copy(header_style.border)

    persons = get_person_rows(ws, name_col, item_col)

    print(f"\n  [表8] 人员数: {len(persons)}")

    for name, start_row, item_rows in persons:
        t7, variant_note = lookup_table7(name, table7_data)
        alerts = []

        # 上午地点
        if "上午地点" in item_rows:
            r = item_rows["上午地点"]
            val = t7["上午出工地点"] if t7 else ""
            ws.cell(row=r, column=fill_col, value=val if val else "")

        # 下午地点
        if "下午地点" in item_rows:
            r = item_rows["下午地点"]
            val = t7["下午出工地点"] if t7 else ""
            ws.cell(row=r, column=fill_col, value=val if val else "")

        # 晚上加班时长
        if "晚上加班时长" in item_rows:
            r = item_rows["晚上加班时长"]
            if t7 and t7["加班时长"]:
                try:
                    ws.cell(row=r, column=fill_col, value=float(t7["加班时长"]))
                except ValueError:
                    ws.cell(row=r, column=fill_col, value=t7["加班时长"])
            else:
                ws.cell(row=r, column=fill_col, value="")

        # 异常提醒
        if variant_note:
            alerts.append(variant_note)
        if t7 is None:
            alerts.append("未在表7中找到出工信息")
        elif t7["类别"] != "公司人员":
            cat_display = t7["类别"] if t7["类别"] else "空"
            alerts.append(f"表7中类别为「{cat_display}」，非公司人员")
        if t7 and t7["异常提醒"]:
            alerts.append(t7["异常提醒"])

        alert_text = "；".join(alerts)
        for r in item_rows.values():
            ws.cell(row=r, column=alert_col, value=alert_text if alert_text else "")

    print(f"  填表完成")
    return ws


def fill_table9(wb, table7_data):
    """填表9外协考勤表1（外协1/临时工）"""
    ws = wb["表9外协考勤表1模板"]

    # 列: A=序号, B=姓名, C=地点, D=出工情况, E=异常提醒(new)
    name_col = 2  # B
    item_col = 3  # C
    fill_col = 4  # D
    alert_col = 5  # E (new)

    # 添加异常提醒表头
    ws.cell(row=1, column=alert_col, value="异常提醒")
    header_style = ws.cell(row=1, column=fill_col)
    alert_header = ws.cell(row=1, column=alert_col)
    if header_style.font:
        alert_header.font = copy(header_style.font)
    if header_style.fill:
        alert_header.fill = copy(header_style.fill)
    if header_style.alignment:
        alert_header.alignment = copy(header_style.alignment)
    if header_style.border:
        alert_header.border = copy(header_style.border)

    persons = get_person_rows(ws, name_col, item_col)

    print(f"\n  [表9] 人员数: {len(persons)}")

    for name, start_row, item_rows in persons:
        t7, variant_note = lookup_table7(name, table7_data)
        alerts = []

        # 上午地点
        if "上午地点" in item_rows:
            r = item_rows["上午地点"]
            val = t7["上午出工地点"] if t7 else ""
            ws.cell(row=r, column=fill_col, value=val if val else "")

        # 下午地点
        if "下午地点" in item_rows:
            r = item_rows["下午地点"]
            val = t7["下午出工地点"] if t7 else ""
            ws.cell(row=r, column=fill_col, value=val if val else "")

        # 加班时长
        if "加班时长" in item_rows:
            r = item_rows["加班时长"]
            if t7 and t7["加班时长"]:
                try:
                    ws.cell(row=r, column=fill_col, value=float(t7["加班时长"]))
                except ValueError:
                    ws.cell(row=r, column=fill_col, value=t7["加班时长"])
            else:
                ws.cell(row=r, column=fill_col, value="")

        # 异常提醒
        if variant_note:
            alerts.append(variant_note)
        if t7 is None:
            alerts.append("未在表7中找到出工信息")
        elif t7["类别"] not in ("外协1", "临时工"):
            cat_display = t7["类别"] if t7["类别"] else "空"
            alerts.append(f"表7中类别为「{cat_display}」，非外协1/临时工")
        if t7 and t7["异常提醒"]:
            alerts.append(t7["异常提醒"])

        alert_text = "；".join(alerts)
        for r in item_rows.values():
            ws.cell(row=r, column=alert_col, value=alert_text if alert_text else "")

    print(f"  填表完成")
    return ws


def fill_table10(wb, table7_data):
    """填表10外协考勤表2（外协2）"""
    ws = wb["表10外协考勤表2模板"]

    # 列: A=序号, B=姓名, C=地点, D=出工情况, E=所属人员, F=异常提醒(new)
    name_col = 2  # B
    item_col = 3  # C
    fill_col = 4  # D
    belong_col = 5  # E
    alert_col = 6  # F (new)

    # 添加异常提醒表头
    ws.cell(row=1, column=alert_col, value="异常提醒")
    header_style = ws.cell(row=1, column=fill_col)
    alert_header = ws.cell(row=1, column=alert_col)
    if header_style.font:
        alert_header.font = copy(header_style.font)
    if header_style.fill:
        alert_header.fill = copy(header_style.fill)
    if header_style.alignment:
        alert_header.alignment = copy(header_style.alignment)
    if header_style.border:
        alert_header.border = copy(header_style.border)

    persons = get_person_rows(ws, name_col, item_col)

    print(f"\n  [表10] 人员数: {len(persons)}")

    for name, start_row, item_rows in persons:
        t7, variant_note = lookup_table7(name, table7_data)
        alerts = []

        # 上午地点
        if "上午地点" in item_rows:
            r = item_rows["上午地点"]
            val = t7["上午出工地点"] if t7 else ""
            ws.cell(row=r, column=fill_col, value=val if val else "")

        # 下午地点
        if "下午地点" in item_rows:
            r = item_rows["下午地点"]
            val = t7["下午出工地点"] if t7 else ""
            ws.cell(row=r, column=fill_col, value=val if val else "")

        # 加班时长
        if "加班时长" in item_rows:
            r = item_rows["加班时长"]
            if t7 and t7["加班时长"]:
                try:
                    ws.cell(row=r, column=fill_col, value=float(t7["加班时长"]))
                except ValueError:
                    ws.cell(row=r, column=fill_col, value=t7["加班时长"])
            else:
                ws.cell(row=r, column=fill_col, value="")

        # 所属人员 (表10独有)
        if t7 and t7["所属人员"]:
            for r in item_rows.values():
                ws.cell(row=r, column=belong_col, value=t7["所属人员"])

        # 异常提醒
        if variant_note:
            alerts.append(variant_note)
        if t7 is None:
            alerts.append("未在表7中找到出工信息")
        elif t7["类别"] != "外协2":
            cat_display = t7["类别"] if t7["类别"] else "空"
            alerts.append(f"表7中类别为「{cat_display}」，非外协2")
        if t7 and t7["异常提醒"]:
            alerts.append(t7["异常提醒"])

        alert_text = "；".join(alerts)
        for r in item_rows.values():
            ws.cell(row=r, column=alert_col, value=alert_text if alert_text else "")

    print(f"  填表完成")
    return ws


def save_sheet_as_file(template_wb, sheet_name, output_path):
    """将模板wb中的指定sheet保存为独立xlsx文件，保留格式和合并单元格"""
    new_wb = Workbook()
    new_wb.remove(new_wb.active)

    ws = template_wb[sheet_name]
    new_ws = new_wb.create_sheet(title=sheet_name)

    # 复制所有单元格的值和样式
    for row in ws.iter_rows():
        for cell in row:
            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.alignment = copy(cell.alignment)
                new_cell.border = copy(cell.border)
                new_cell.number_format = cell.number_format

    # 复制列宽
    for col_letter, col_dim in ws.column_dimensions.items():
        new_ws.column_dimensions[col_letter].width = col_dim.width

    # 复制行高
    for row_num, row_dim in ws.row_dimensions.items():
        new_ws.row_dimensions[row_num].height = row_dim.height

    # 复制合并单元格
    for merge_range in ws.merged_cells.ranges:
        new_ws.merge_cells(str(merge_range))

    new_wb.save(output_path)
    print(f"  已保存: {output_path}")


def get_table8_9_10():
    """
    从表7填至 表8/表9/表10 最终考勤表。

    - 07_表7 类别=公司人员 → 08_表8工程考勤表
    - 07_表7 类别=外协1/临时工 → 09_表9外协考勤表1
    - 07_表7 类别=外协2 → 10_表10外协考勤表2

    保持模板原有姓名顺序，末尾新增"异常提醒"列。
    """
    print("=" * 50)
    print("  填表8/表9/表10 从07_表7出工地点及时长统计表")
    print("=" * 50)

    # 1. 读取表7
    table7_data = read_table7()
    company_count = sum(1 for v in table7_data.values() if v["类别"] == "公司人员")
    waixie1_count = sum(
        1 for v in table7_data.values() if v["类别"] in ("外协1", "临时工")
    )
    waixie2_count = sum(1 for v in table7_data.values() if v["类别"] == "外协2")
    unknown_count = len(table7_data) - company_count - waixie1_count - waixie2_count
    print(f"表7总人数: {len(table7_data)}")
    print(
        f"  公司人员: {company_count}  |  外协1/临时工: {waixie1_count}  |  外协2: {waixie2_count}  |  未分类: {unknown_count}"
    )

    # 2. 打开模板文件
    wb = load_workbook(TEMPLATE_FILE)
    print(f"\n加载模板: {TEMPLATE_FILE}")

    # 3. 依次填表
    fill_table8(wb, table7_data)
    fill_table9(wb, table7_data)
    fill_table10(wb, table7_data)

    # 4. 保存为"一个文件三个子表"（表8/表9/表10，与输入模板同构；不修改模板文件本身）
    sheet_rename = {
        "表8工程考勤表模板": "表8工程考勤表",
        "表9外协考勤表1模板": "表9外协考勤表1",
        "表10外协考勤表2模板": "表10外协考勤表2",
    }
    for old, new in sheet_rename.items():
        if old in wb.sheetnames:
            wb[old].title = new
    # 删除模板自带的空 sheet（如 Sheet1）
    for ws in list(wb.worksheets):
        if ws.max_row <= 1 and ws.max_column <= 1 and ws["A1"].value in (None, ""):
            wb.remove(ws)
    combined_file = OUTPUT_DIR / "08_09_10_最终考勤表.xlsx"
    wb.save(combined_file)
    print(f"\n生成最终考勤表（一个文件三个子表）: {combined_file}")
    print(f"  子表: {wb.sheetnames}")

    wb.close()
    print(f"\n完成!")
    print("=" * 50)
