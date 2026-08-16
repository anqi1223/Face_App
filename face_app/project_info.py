"""生成 000_项目信息表。

输入:
  input/04_工地项目统计.xlsx   (项目名称 | 简称, 大项目表)
  input/06_周工作安排表.xlsx    (本周安排, 标题含日期范围)
  input/07_周工作计划表.xlsx    (按天计划, 只取日期范围内)
输出:
  output/000_项目信息表.xlsx    (工作安排简称 | 项目名称 | 出勤统计简称, 与原04同构)
"""

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent.parent
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
PROJECT_STAT_FILE = INPUT_DIR / "04_工地项目统计.xlsx"
WEEKLY_ARRANGE_FILE = INPUT_DIR / "06_周工作安排表.xlsx"
WEEKLY_PLAN_FILE = INPUT_DIR / "07_周工作计划表.xlsx"
PROJECT_INFO_FILE = OUTPUT_DIR / "000_项目信息表.xlsx"
UPDATE_STAT_FILE = OUTPUT_DIR / "0000_工地项目统计更新.xlsx"

OUTPUT_COLS = ["工作安排简称", "项目名称", "出勤统计简称"]


# ============================================================
# 工具函数
# ============================================================
def _norm(s) -> str:
    """规范化项目名：去空格、统一大小写/kV、去尾部"项目"。"""
    if s is None:
        return ""
    s = re.sub(r"\s+", "", str(s).strip())
    s = s.lower().replace("kv", "kV")
    if s.endswith("项目"):
        s = s[:-2]
    return s


def _substation_short(s) -> str:
    """从变电站名提取基础简称：'220kV震泽变' → '震泽'。"""
    s = re.sub(r"\s+", "", str(s or ""))
    s = re.sub(r"\d+kv", "", s, flags=re.IGNORECASE)
    s = s.replace("变电站", "").replace("变", "")
    return s.strip()


def _next_short(base: str, used: set) -> str:
    """简称递增：已有'震泽'→'震泽1'；已有'震泽1'→'震泽2'。"""
    base = base or "新增"
    if base not in used:
        return base
    i = 1
    while f"{base}{i}" in used:
        i += 1
    return f"{base}{i}"


def _trailing_short(orig: str) -> str:
    """提取项目名末尾的变电站简称：'国网…蓄电池改造-震泽' → '震泽'。"""
    m = re.search(r"[-—]([^-—]+)$", str(orig or "").strip())
    return m.group(1).strip() if m else ""


def _parse_week_range(title: str):
    """从 06 标题解析日期范围，如 '2026年8月9日-8月15日周工作安排表'。
    返回 (起始月日, 结束月日) 均为 (月, 日)。支持跨月 '8月30日-9月5日'。"""
    m = re.search(
        r"(\d{1,2})月(\d{1,2})日\s*-\s*(\d{1,2})月(\d{1,2})日", str(title or "")
    )
    if not m:
        return None
    return (int(m.group(1)), int(m.group(2))), (int(m.group(3)), int(m.group(4)))


def _in_range(md, start, end) -> bool:
    """(月,日) 是否落在 [start, end] 区间内（考虑跨月）。"""
    start_val, end_val = start[0] * 100 + start[1], end[0] * 100 + end[1]
    val = md[0] * 100 + md[1]
    if start_val <= end_val:
        return start_val <= val <= end_val
    # 跨月：如 8/30 - 9/5
    return val >= start_val or val <= end_val


# ============================================================
# 读表
# ============================================================
def read_project_stat():
    """读取 04_工地项目统计。

    注意：同一个项目名可能有多条（等变电站项目不同变电站 → 不同简称），
    因此按规范化名存「列表」，每条 (简称, 原始项目名称)。
    返回 ({norm_name: [(简称, 原始名), ...]}, 全部简称集合)。
    """
    if not PROJECT_STAT_FILE.exists():
        return {}, set()
    df = pd.read_excel(PROJECT_STAT_FILE, sheet_name="工地项目统计")
    df.columns = [str(c).strip() for c in df.columns]
    lookup = {}
    used = set()
    for _, r in df.iterrows():
        name = str(r.get("项目名称", "")).strip()
        short = str(r.get("简称", "")).strip()
        if name:
            lookup.setdefault(_norm(name), []).append((short, name))
        if short:
            used.add(short)
    return lookup, used


def read_weekly_arrange():
    """读取 06_周工作安排表数据行。返回 (period, rows)。
    rows: [{序号, 项目名称, 项目单位, 单位简称, 所涉变电站, 工作安排}]"""
    wb = load_workbook(WEEKLY_ARRANGE_FILE, data_only=True)
    ws = wb.active
    title = ws.cell(row=1, column=1).value
    period = _parse_week_range(title)
    rows = []
    for r in range(3, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value  # B=项目名称
        if not name or str(name).strip() == "":
            continue
        unit = ws.cell(row=r, column=3).value  # C=项目单位
        substation = ws.cell(row=r, column=4).value  # D=所涉变电站
        arrange = ws.cell(row=r, column=5).value  # E=工作安排
        # 从项目单位提取简称：'…（简称：河南平高）'
        unit_short = ""
        um = re.search(r"简称[:：]\s*([^）)]+)", str(unit or ""))
        if um:
            unit_short = um.group(1).strip()
        rows.append({
            "序号": ws.cell(row=r, column=1).value,
            "项目名称": str(name).strip(),
            "项目单位": str(unit).strip() if unit else "",
            "单位简称": unit_short,
            "所涉变电站": str(substation).strip() if substation else "",
            "工作安排": str(arrange).strip() if arrange else "",
        })
    wb.close()
    return period, rows


def read_weekly_plan(start, end):
    """读取 07_周工作计划表，取日期范围内的 项目(B列) 与 工作内容(H列)。
    返回 [(项目, 工作内容), ...] 与 去重后的项目集合。"""
    wb = load_workbook(WEEKLY_PLAN_FILE, data_only=True)
    entries = []
    day = None
    header_done = False
    for ws in wb.worksheets:
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            dm = re.search(r"^(\d{1,2})月(\d{1,2})日工作安排", str(a or ""))
            if dm:
                md = (int(dm.group(1)), int(dm.group(2)))
                day = md if _in_range(md, start, end) else None
                header_done = False
                continue
            if day is None:
                continue
            # 表头行（序号/项目/...）跳过
            if not header_done and a is not None and str(a).strip() == "序号":
                header_done = True
                continue
            if header_done:
                proj = ws.cell(row=r, column=2).value  # B=项目
                if proj and str(proj).strip():
                    content = ws.cell(row=r, column=8).value  # H=工作内容
                    entries.append((str(proj).strip(),
                                    str(content).strip() if content else ""))
    wb.close()
    return entries, list(dict.fromkeys(p for p, _ in entries))


# ============================================================
# 生成计划
# ============================================================
def generate_plan():
    """生成 000 的确认计划。返回 {period, auto, review}，不写文件。
    auto: [{工作安排简称, 项目名称, 出勤统计简称}] 已自动确定
    review: [{row_no, 项目名称, 所涉变电站, 单位简称, 工作安排,
              work_short, work_short_options, attendance_short, short_options,
              mode(fuzzy|none|new), suggest_name}] 需人工确认
    """
    period, arrange_rows = read_weekly_arrange()
    if not period:
        raise ValueError("无法从 06_周工作安排表 标题解析日期范围")

    stat_lookup, stat_shorts = read_project_stat()
    plan_entries, plan_projects = read_weekly_plan(*period)

    # 07 项目集合（规范化）
    plan_proj_norm = {_norm(p) for p in plan_projects}

    auto = []
    review = []
    # 04 已有简称 + 0000 更新表已有简称，都计入避免重复（简称编号跨周延续）
    used_shorts = set(stat_shorts) | read_update_shorts()
    rid = 0  # 待确认项唯一序号（06 的"序号"列可能有重复）

    def _find_candidates(name_norm, name):
        """在 04 中找相似候选，返回 [(简称, 项目名称), ...]"""
        cands = []
        for k, entries in stat_lookup.items():
            for (short, orig) in entries:
                if not short:
                    continue
                if name_norm == k:
                    return [(short, orig)]  # 完全一致（取第一条）
                if name_norm and (name_norm in k or k in name_norm):
                    cands.append((short, orig))
        # 按长度相似排序（更长的前缀/包含更贴近）
        cands.sort(key=lambda c: len(c[1]), reverse=True)
        return cands[:8]

    for item in arrange_rows:
        name = item["项目名称"]
        substation = item["所涉变电站"]
        name_norm = _norm(name)
        base = _substation_short(substation) or name
        is_etc = ("等变电站" in name) or ("等" in name and "座变电站" in name)
        is_special = (not re.search(r"\d+kv", name, re.IGNORECASE)) and ("变" not in name)

        # --- 工作安排简称：用所涉变电站匹配 07 项目列 ---
        work_short = substation
        work_opts = []
        sub_norm = _norm(substation)
        matched = [p for p in plan_projects if _norm(p) == sub_norm]
        if not matched and sub_norm:
            matched = [p for p in plan_projects if sub_norm in _norm(p) or _norm(p) in sub_norm]
        if matched:
            work_short = matched[0]
            work_opts = list(dict.fromkeys(matched + [work_short]))[:6]
        else:
            work_opts = [substation]
        if is_special and substation and work_short == substation:
            # 特殊地点（公司/仓库/项目部等）：工作安排简称优先用名称本身
            work_short = name
            work_opts = list(dict.fromkeys([name] + work_opts))[:6]

        # --- 出勤统计简称 ---
        cands = _find_candidates(name_norm, name)

        def _rel(base):
            """04 中与某变电站基础简称相关的项目说明（简称==base 或 base+N）。"""
            out = []
            for _k, entries in stat_lookup.items():
                for (short, orig) in entries:
                    if short and (short == base or re.fullmatch(re.escape(base) + r"\d*", short)):
                        out.append(f"{orig}（简称：{short}）")
            return out

        def _mk_review(mode, default_short, short_opts, suggest_name, reason):
            nonlocal rid
            rid += 1
            review.append({
                "rid": rid, "row_no": item["序号"], "项目名称": name,
                "所涉变电站": substation,
                "单位简称": item["单位简称"], "工作安排": item["工作安排"],
                "work_short": work_short, "work_short_options": work_opts,
                "attendance_short": default_short,
                "short_options": short_opts,
                "mode": mode, "suggest_name": suggest_name, "reason": reason,
            })

        if is_etc:
            # 等变电站项目：04 里同伞形名可有多条，简称对应不同变电站（如 南施/南施1/…/南施10）。
            # 按 简称 是否对应所涉变电站匹配（优先精确简称），匹配到就用 04 原有项目名+简称；否则新增行
            specific = []
            for k, entries in stat_lookup.items():
                if not (name_norm in k or k in name_norm):
                    continue
                for (short, orig) in entries:
                    if not short:
                        continue
                    if short == base or re.fullmatch(re.escape(base) + r"\d*", short):
                        specific.append((short, orig))
            specific.sort(key=lambda x: x[0] != base)  # 精确简称优先
            if specific:
                short, orig = specific[0][0], specific[0][1]
                used_shorts.add(short)
                auto.append({"工作安排简称": work_short, "项目名称": orig,
                             "出勤统计简称": short})
            else:
                new_short = _next_short(base, used_shorts)
                used_shorts.add(new_short)
                # 该伞形项目在 04 里的条目说明
                umbrella = []
                for k, entries in stat_lookup.items():
                    if name_norm in k or k in name_norm:
                        for (short, orig) in entries:
                            if short:
                                umbrella.append(f"{orig}（简称：{short}）")
                reason = (
                    f"在04_工地项目统计中「{substation}」在该伞形项目下无对应条目"
                    + (f"（该伞形项目仅有：『{'; '.join(umbrella[:4])}』）" if umbrella else "")
                    + "，需新增行"
                )
                _mk_review("new", new_short, [new_short, "新增"],
                           f"{name}—{base}", reason)
            continue

        if is_special:
            # 特殊地点：建议简称 = 名称本身
            reason = f"在04_工地项目统计中未找到「{name}」相关项目，按名称本身作为出勤统计简称处理"
            _mk_review("none", name, [name, "新增"], name, reason)
            continue

        if cands and _norm(cands[0][1]) == name_norm:
            # 完全一致 → 自动，用 04 里原有的项目名
            short, orig = cands[0][0], cands[0][1]
            used_shorts.add(short)
            auto.append({"工作安排简称": work_short, "项目名称": orig,
                         "出勤统计简称": short})
            continue

        if cands:
            # 相近候选 → 确认，默认取最相近（项目名用 04 里原有的）
            rel = _rel(base)
            if rel:
                reason = (
                    f"在04_工地项目统计中「{substation}」变电站仅有项目：『{'; '.join(rel[:3])}』；"
                    f"当前项目「{name}」无法直接匹配（相近项目「{cands[0][1]}」简称{cands[0][0]}），请人工确认"
                )
            else:
                reason = (
                    f"在04_工地项目统计中未找到「{substation}」相关项目；"
                    f"当前项目「{name}」无法直接匹配（相近项目「{cands[0][1]}」简称{cands[0][0]}），请人工确认"
                )
            _mk_review("fuzzy", cands[0][0], [s for s, _ in cands] + ["新增"],
                       cands[0][1], reason)
        else:
            # 无匹配 → 建议新增
            new_short = _next_short(base, used_shorts)
            used_shorts.add(new_short)
            rel = _rel(base)
            if rel:
                reason = (
                    f"在04_工地项目统计中「{substation}」变电站仅有项目：『{'; '.join(rel[:3])}』；"
                    f"当前项目「{name}」无法直接匹配，需新增"
                )
            else:
                reason = (
                    f"在04_工地项目统计中未找到「{substation}」相关项目；"
                    f"当前项目「{name}」无法直接匹配，需新增"
                )
            _mk_review("none", new_short, ["新增", new_short], name, reason)

    return {
        "period": f"{period[0][0]}/{period[0][1]} - {period[1][0]}/{period[1][1]}",
        "auto": auto,
        "review": review,
    }


def write_000(rows, highlight_idxs=None):
    """写入 output/000_项目信息表.xlsx。
    rows: [{工作安排简称,项目名称,出勤统计简称}]
    highlight_idxs: 新增行（行下标）用单元格黄色底色突出显示。
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows, columns=OUTPUT_COLS)
    df.to_excel(PROJECT_INFO_FILE, index=False)
    if highlight_idxs:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        wb = load_workbook(PROJECT_INFO_FILE)
        ws = wb.active
        fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for idx in highlight_idxs:
            r = idx + 2  # 第1行是表头
            for c in range(1, len(OUTPUT_COLS) + 1):
                ws.cell(row=r, column=c).fill = fill
        wb.save(PROJECT_INFO_FILE)
        wb.close()
    return str(PROJECT_INFO_FILE), len(df)


def read_update_shorts() -> set:
    """读取 0000_工地项目统计更新表 中已有的简称（用于跨周编号延续）。"""
    if not UPDATE_STAT_FILE.exists():
        return set()
    try:
        df = pd.read_excel(UPDATE_STAT_FILE)
        return set(df["简称"].astype(str).str.strip()) if "简称" in df.columns else set()
    except Exception:
        return set()


def update_project_stat_0000(new_items, period):
    """把 000 中新增的项目行插入 0000_工地项目统计更新表。

    new_items: [{项目名称, 出勤统计简称}]（000 中 mode=new/none 的行）
    放在同一变电站（简称前缀）原有行下方（如 石牌9 放在 石牌8 下面）；
    新增行黄色高亮，并记录"新增周期"列。
    返回 (路径, 总行数)。
    """
    cols = ["项目名称", "简称", "新增周期"]
    if UPDATE_STAT_FILE.exists():
        df = pd.read_excel(UPDATE_STAT_FILE)
        for c in cols:
            if c not in df.columns:
                df[c] = ""
        rows = df[cols].astype(object).values.tolist()
    else:
        # 首次生成：以 04_工地项目统计 为底，新行才能插在对应变电站原有行下方
        base = pd.read_excel(PROJECT_STAT_FILE, sheet_name="工地项目统计")
        base.columns = ["项目名称", "简称"]
        base = base.copy()
        base["新增周期"] = ""
        rows = base[cols].astype(object).values.tolist()

    def base_of(s):
        return re.sub(r"\d+$", "", str(s).strip())

    inserted = 0
    for item in new_items:
        name = str(item.get("项目名称", "")).strip()
        short = str(item.get("出勤统计简称", "")).strip()
        if not name or not short:
            continue
        if any(str(r[1]).strip() == short for r in rows):
            continue  # 已存在则跳过
        base = base_of(short)
        # 定位：新简称 base+N 插在 base+(N-1) 下面（如 石牌9→石牌8 下、寒山5→寒山4 下）
        m = re.fullmatch(re.escape(base) + r"(\d+)", short)
        my_num = int(m.group(1)) if m else 0
        target = base if my_num == 0 else base + str(my_num - 1)
        insert_after = -1
        for i, r in enumerate(rows):
            if str(r[1]).strip() == target:
                insert_after = i
                break
        if insert_after < 0:
            # 回退：最后一个同 base 组且编号 < my_num 的行
            for i, r in enumerate(rows):
                s = str(r[1]).strip()
                if s == base:
                    if my_num > 0:
                        insert_after = i
                else:
                    mm = re.fullmatch(re.escape(base) + r"(\d+)", s)
                    if mm and int(mm.group(1)) < my_num:
                        insert_after = i
        rows.insert(insert_after + 1, [name, short, period])
        inserted += 1

    out = pd.DataFrame(rows, columns=cols)
    out.to_excel(UPDATE_STAT_FILE, index=False)
    # 高亮"新增周期"非空的行
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    wb = load_workbook(UPDATE_STAT_FILE)
    ws = wb.active
    fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=3).value:
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = fill
    wb.save(UPDATE_STAT_FILE)
    wb.close()
    return str(UPDATE_STAT_FILE), inserted
